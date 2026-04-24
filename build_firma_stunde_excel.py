# -*- coding: utf-8 -*-
"""
PDF puantaj tablosundan profesyonel Excel çalışma kitabı üretir.

Çalıştırma:  python build_firma_stunde_excel.py

Çıktı:       firma_stunde_calisma_sayfasi.xlsx (TALIP + isteğe bağlı İndirilenler)
"""
from __future__ import annotations

import logging
import re
import sys
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# --- Sabitler ----------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
OUT_NAME = "firma_stunde_calisma_sayfasi.xlsx"
OUT_PATH = SCRIPT_DIR / OUT_NAME
OUT_PATH_DOWNLOADS = Path.home() / "Downloads" / OUT_NAME

PDF_CANDIDATES = [
    # Cursor workspace (güncel ham PDF — öncelik)
    Path(
        r"c:\Users\issan\AppData\Roaming\Cursor\User\workspaceStorage"
        r"\31493f056ac4fc2dcff79db6af47f891\pdfs\5f21082c-1cd1-4344-9932-530b39f02c87"
        r"\Firma stunde.pdf"
    ),
    SCRIPT_DIR / "Firma stunde.pdf",
    SCRIPT_DIR / "firma_stunde.pdf",
    Path(
        r"c:\Users\issan\AppData\Roaming\Cursor\User\workspaceStorage"
        r"\31493f056ac4fc2dcff79db6af47f891\pdfs\eae826af-a0fa-4223-b1bd-d7c258625211"
        r"\Firma stunde.pdf"
    ),
    Path.home() / "Downloads" / "Firma stunde.pdf",
]

LOG = logging.getLogger("firma_stunde")

THIN = Side(style="thin", color="FF333333")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="FF2F5597")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=10)
DATA_FONT = Font(size=9)
TITLE_FONT = Font(bold=True, size=11)
NAME_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")

CF_OK = PatternFill("solid", fgColor="FFC6EFCE")
CF_HATA = PatternFill("solid", fgColor="FFFFC7CE")
CF_PDF_HATA = PatternFill("solid", fgColor="FFFCE4D6")

TOLERANCE = 0.02


# --- Normalizasyon -----------------------------------------------------------


def normalize_cell_display(raw: Any) -> Any:
    if raw is None:
        return None
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return raw
    s = str(raw).strip()
    return s if s else None


def parse_eu_number(s: str | None) -> float | None:
    if s is None:
        return None
    t = re.sub(r"\s+", "", str(s).strip()).replace("\u00a0", "")
    if t in ("", "-", "#WERT!", "#VALUE!", "#REF!", "#DIV/0!"):
        return None
    t = re.sub(r"^€+", "", t, flags=re.I)
    if "," in t:
        t = t.replace(".", "").replace(",", ".")
    try:
        return float(Decimal(t))
    except (InvalidOperation, ValueError):
        return None


def parse_currency_eu(s: str | None) -> float | None:
    if s is None:
        return None
    t = re.sub(r"\s+", "", str(s).strip()).replace("\u00a0", "")
    if not t or t in ("-", "#WERT!", "#VALUE!"):
        return None
    t = re.sub(r"€", "", t, flags=re.I)
    if "," in t:
        t = t.replace(".", "").replace(",", ".")
    try:
        return float(Decimal(t))
    except (InvalidOperation, ValueError):
        return None


def parse_saat_ucreti_cell(s: str | None) -> float:
    n = parse_currency_eu(s)
    if n is not None:
        return n
    if not s:
        return 1.0
    nums = re.findall(r"\d+[.,]?\d*", str(s).replace(" ", ""))
    for x in nums:
        try:
            return float(x.replace(",", "."))
        except ValueError:
            continue
    return 1.0


def is_pdf_formula_error_text(s: str | None) -> bool:
    if not s:
        return False
    return str(s).strip().upper() in ("#WERT!", "#VALUE!", "#REF!", "#DIV/0!")


def day_cell_for_working_sheet(raw: Any) -> Any:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    if s == "-":
        return "-"
    n = parse_eu_number(s)
    if n is not None:
        return n
    return s


# --- PDF -------------------------------------------------------------------


def resolve_pdf_path() -> Path:
    for p in PDF_CANDIDATES:
        try:
            if p.exists() and p.is_file():
                LOG.info("PDF bulundu: %s", p)
                return p
        except OSError:
            continue
    raise FileNotFoundError(
        "PDF bulunamadı. 'Firma stunde.pdf' dosyasını şu klasöre koyun: " + str(SCRIPT_DIR)
    )


def extract_table_and_text(path: Path) -> tuple[list[list[Any]], list[str]]:
    """PDF’i açar, tabloyu ve sayfa metnini döndürür (with ile kapanır)."""
    with pdfplumber.open(path) as pdf:
        page = pdf.pages[0]
        tables = page.extract_tables() or []
        if not tables:
            raise ValueError("PDF sayfasında tablo bulunamadı.")
        text_lines = extract_page_text_lines(page)
        return tables[0], text_lines


def extract_page_text_lines(page) -> list[str]:
    return [ln.rstrip() for ln in (page.extract_text() or "").splitlines()]


# --- Sütun haritası --------------------------------------------------------


@dataclass
class ColumnMap:
    col_sira: int = 0
    col_name: int = 1
    day_cols: list[int] = field(default_factory=list)
    col_stunde: int | None = None
    col_saat: int | None = None
    col_toplam: int | None = None


def _hcell(x: Any) -> str:
    if x is None:
        return ""
    return re.sub(r"\s+", " ", str(x).strip()).lower()


def detect_header_row_index(table: list[list[Any]]) -> int:
    for i, row in enumerate(table):
        if not row:
            continue
        cells = [_hcell(c) for c in row]
        joined = " ".join(cells)
        if "adi" in joined and "soyadi" in joined:
            return i
        if "soyadi" in joined:
            return i
    for i, row in enumerate(table):
        joined = _hcell(" ".join(str(c) for c in row if c is not None))
        if "stunde" in joined and "toplam" in joined:
            return i
    return max(0, min(2, len(table) - 1))


def build_column_map(header_row: list[Any]) -> ColumnMap:
    cm = ColumnMap()
    if not header_row:
        return cm
    texts = [_hcell(c) for c in header_row]
    n = len(header_row)

    for j, t in enumerate(texts):
        if "soyadi" in t or t == "ad":
            cm.col_name = j
            break
    for j in range(cm.col_name):
        if j != cm.col_name:
            cm.col_sira = j
            break

    day_pairs: list[tuple[int, int]] = []
    for j, t in enumerate(texts):
        if re.fullmatch(r"\d{1,2}", t or ""):
            num = int(t)
            if 1 <= num <= 31:
                day_pairs.append((num, j))
    day_pairs.sort(key=lambda x: x[0])
    if len(day_pairs) >= 28:
        cm.day_cols = [idx for _, idx in day_pairs[:31]]
    else:
        start = cm.col_name + 1
        while start < n and header_row[start] is None:
            start += 1
        cm.day_cols = [start + k for k in range(31) if start + k < n]

    for j, t in enumerate(texts):
        if "stunde" in t:
            cm.col_stunde = j
        if "ücret" in t or "ucret" in t:
            cm.col_saat = j
        if t == "toplam" or (t.startswith("toplam") and "çalışan" not in t):
            cm.col_toplam = j

    if cm.col_stunde is None and cm.day_cols:
        cm.col_stunde = max(cm.day_cols) + 1
    if cm.col_saat is None and cm.col_stunde is not None:
        cm.col_saat = cm.col_stunde + 1
    if cm.col_toplam is None and cm.col_saat is not None:
        cm.col_toplam = cm.col_saat + 1

    return cm


def row_looks_like_footer(row: list[Any]) -> bool:
    if not row or row[0] is None:
        return False
    s = str(row[0]).upper()
    return "TOPLAM" in s and ("ÇALIŞAN" in s or "CALISAN" in s)


def row_looks_like_employee(row: list[Any], cm: ColumnMap) -> bool:
    if not row or len(row) <= cm.col_name:
        return False
    if row_looks_like_footer(row):
        return False
    first = row[cm.col_sira] if cm.col_sira < len(row) else None
    name = row[cm.col_name] if cm.col_name < len(row) else None
    if name is None or str(name).strip() == "":
        return False
    fs = str(first).strip() if first is not None else ""
    if fs.isdigit():
        return True
    if re.match(r"^\d+$", fs):
        return True
    return False


@dataclass
class EmployeeRow:
    sira: Any
    ad: str
    days: list[Any]
    pdf_stunde_raw: Any
    pdf_saat_raw: Any
    pdf_toplam_raw: Any


def parse_employee_rows(
    table: list[list[Any]], header_idx: int, cm: ColumnMap
) -> tuple[list[EmployeeRow], int | None]:
    """Çalışan satırları ve (varsa) footer satır indeksi."""
    out: list[EmployeeRow] = []
    footer_idx: int | None = None
    for i in range(header_idx + 1, len(table)):
        row = table[i]
        if not row:
            continue
        if row_looks_like_footer(row):
            footer_idx = i
            break
        if not row_looks_like_employee(row, cm):
            continue
        days: list[Any] = []
        for dc in cm.day_cols:
            days.append(row[dc] if dc < len(row) else None)
        st_r = row[cm.col_stunde] if cm.col_stunde is not None and cm.col_stunde < len(row) else None
        sa_r = row[cm.col_saat] if cm.col_saat is not None and cm.col_saat < len(row) else None
        to_r = row[cm.col_toplam] if cm.col_toplam is not None and cm.col_toplam < len(row) else None
        out.append(
            EmployeeRow(
                sira=row[cm.col_sira] if cm.col_sira < len(row) else len(out) + 1,
                ad=str(row[cm.col_name] or "").strip(),
                days=days,
                pdf_stunde_raw=st_r,
                pdf_saat_raw=sa_r,
                pdf_toplam_raw=to_r,
            )
        )
    return out, footer_idx


def parse_footer_row(table: list[list[Any]], footer_idx: int | None) -> list[Any] | None:
    if footer_idx is None or footer_idx >= len(table):
        return None
    return table[footer_idx]


def parse_legend_lines(table: list[list[Any]], footer_idx: int | None, text_lines: list[str]) -> list[str]:
    """Tablo sonrası lejant / isim listesi için en iyi çaba (yinelenen satırlar sınırlı)."""
    legend: list[str] = []
    seen: set[str] = set()
    start = (footer_idx + 1) if footer_idx is not None else len(table)
    for i in range(start, len(table)):
        row = table[i]
        if not row:
            continue
        parts = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if parts:
            line = " | ".join(parts)
            if line not in seen:
                seen.add(line)
                legend.append(line)
    markers = ("senelik", "yarım", "çıkış", "dış görev", "nisan ayı", "toplam çalışan", "işe devam")
    capture = False
    for ln in text_lines:
        low = ln.lower()
        if any(m in low for m in markers):
            capture = True
        if capture and ln.strip():
            s = ln.strip()
            if s not in seen:
                seen.add(s)
                legend.append(s)
        if len(legend) >= 120:
            break
    return legend[:120]


# --- Sayfa oluşturma --------------------------------------------------------


def apply_print_a4_landscape(ws, last_row: int, last_col: int) -> None:
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToWidth = 1
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:1"
    lc = get_column_letter(last_col)
    ws.print_area = f"A1:{lc}{last_row}"
    ws.page_margins = PageMargins(left=0.45, right=0.45, top=0.55, bottom=0.55)


def style_header_row(ws, row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER_ALL
        cell.alignment = CENTER if c > 2 else NAME_ALIGN


def autofit_columns(ws, max_col: int, min_w: float = 2.5, max_w: float = 40) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        m = min_w
        for row in range(1, min(ws.max_row, 200) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            L = len(str(v))
            m = max(m, min(max_w, L * 0.9 + 1.5))
        ws.column_dimensions[letter].width = m


def build_pdf_sheet(wb: Workbook, table: list[list[Any]], legend_lines: list[str]) -> None:
    ws = wb.create_sheet("PDF_AKTARIM", 0)
    ws.cell(1, 1, "Nisan AYI (PDF ham aktarım)").font = TITLE_FONT

    start_r = 3
    max_c = 1
    for ri, row in enumerate(table):
        r = start_r + ri
        if not row:
            continue
        for ci, val in enumerate(row):
            c = ci + 1
            max_c = max(max_c, c)
            cell = ws.cell(row=r, column=c, value=normalize_cell_display(val))
            cell.font = DATA_FONT
            cell.border = BORDER_ALL
            cell.alignment = NAME_ALIGN if c <= 2 else CENTER

    leg_r = start_r + len(table) + 2
    ws.cell(leg_r, 1, "Lejant / ek metin (PDF)").font = TITLE_FONT
    for i, line in enumerate(legend_lines):
        cell = ws.cell(leg_r + 1 + i, 1, value=line)
        cell.font = DATA_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(
            start_row=leg_r + 1 + i,
            start_column=1,
            end_row=leg_r + 1 + i,
            end_column=min(max_c, 12),
        )

    last_r = leg_r + max(len(legend_lines), 0) + 2
    apply_print_a4_landscape(ws, last_r, max(max_c, 8))
    ws.freeze_panes = ws.cell(start_r + 1, 3)


def build_working_sheet(
    wb: Workbook,
    employees: list[EmployeeRow],
    footer_pdf: list[Any] | None,
    pdf_genel_stunde: float | None,
) -> None:
    ws = wb.create_sheet("CALISMA_SAYFASI")
    day_start = 3
    col_st = day_start + 31
    col_w = col_st + 1
    col_t = col_w + 1
    col_ps = col_t + 1
    col_pt = col_ps + 1
    col_k = col_pt + 1

    headers = (
        ["Sıra No", "Ad Soyad"]
        + [str(d) for d in range(1, 32)]
        + [
            "Stunden (Hesaplanan)",
            "Saat Ücreti",
            "Toplam (Hesaplanan)",
            "PDF Stunden",
            "PDF Toplam",
            "Kontrol",
        ]
    )
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, value=h)

    style_header_row(ws, 1, len(headers))
    c0 = get_column_letter(day_start)
    c1 = get_column_letter(day_start + 30)

    for i, emp in enumerate(employees, start=2):
        ws.cell(i, 1, value=emp.sira)
        ws.cell(i, 2, value=emp.ad)
        for j, raw in enumerate(emp.days):
            ws.cell(i, day_start + j, value=day_cell_for_working_sheet(raw))

        ws.cell(i, col_st, value=f"=SUM({c0}{i}:{c1}{i})")
        wage = parse_saat_ucreti_cell(
            str(emp.pdf_saat_raw) if emp.pdf_saat_raw is not None else None
        )
        ws.cell(i, col_w, value=wage)

        ls = get_column_letter(col_st)
        lw = get_column_letter(col_w)
        ws.cell(i, col_t, value=f"={ls}{i}*{lw}{i}")

        ps = parse_eu_number(str(emp.pdf_stunde_raw).strip() if emp.pdf_stunde_raw else None)
        if ps is not None:
            ws.cell(i, col_ps, value=ps)
        else:
            ws.cell(i, col_ps, value=str(emp.pdf_stunde_raw or "").strip() or None)
        if isinstance(ws.cell(i, col_ps).value, (int, float)):
            ws.cell(i, col_ps).number_format = "#,##0.0"
        else:
            ws.cell(i, col_ps).number_format = "General"

        pt_raw = str(emp.pdf_toplam_raw).strip() if emp.pdf_toplam_raw else ""
        pt_cell = ws.cell(i, col_pt)
        if is_pdf_formula_error_text(pt_raw):
            pt_cell.value = "#WERT!" if "WERT" in pt_raw.upper() else "#VALUE!"
        else:
            pv = parse_currency_eu(pt_raw)
            if pv is None:
                pv = parse_eu_number(pt_raw)
            if pv is not None:
                pt_cell.value = pv
                pt_cell.number_format = '#,##0.00 [$€-40C]'
            else:
                pt_cell.value = pt_raw or None

        lk = get_column_letter(col_k)
        lt = get_column_letter(col_t)
        lps = get_column_letter(col_ps)
        lpt = get_column_letter(col_pt)
        tol = "0.02"
        # PDF_HATA | OK | HATA — PDF Stunden sayıysa saat farkı da kontrol edilir
        ws.cell(
            i,
            col_k,
            value=(
                f'=IF(OR({lpt}{i}="#WERT!",{lpt}{i}="#VALUE!"),"PDF_HATA",'
                f'IF(NOT(ISNUMBER({lpt}{i})),"HATA",'
                f'IF(AND(ISNUMBER({lps}{i}),ABS({ls}{i}-{lps}{i})>={tol}),"HATA",'
                f'IF(ABS({lt}{i}-{lpt}{i})>={tol},"HATA","OK"))))'
            ),
        )

        for c in range(1, col_k + 1):
            cell = ws.cell(i, c)
            cell.font = DATA_FONT
            cell.border = BORDER_ALL
            if c == 2:
                cell.alignment = NAME_ALIGN
            elif day_start <= c < col_st:
                cell.alignment = CENTER
                v = cell.value
                if isinstance(v, (int, float)):
                    cell.number_format = "#,##0.0"
            else:
                cell.alignment = CENTER
        ws.cell(i, col_st).number_format = "#,##0.0"
        ws.cell(i, col_w).number_format = '#,##0.00 [$€-40C]'
        ws.cell(i, col_t).number_format = '#,##0.00 [$€-40C]'

    last_data = 1 + len(employees)
    foot_title_r = last_data + 2
    ws.cell(foot_title_r, 1, "PDF GENEL TOPLAM — ham footer (PDF’teki özet satırı)").font = TITLE_FONT

    r = foot_title_r + 1
    if footer_pdf:
        for ci, val in enumerate(footer_pdf):
            cell = ws.cell(r, ci + 1, value=normalize_cell_display(val))
            cell.font = DATA_FONT
            cell.border = BORDER_ALL
            cell.alignment = CENTER
        r += 1
    if pdf_genel_stunde is not None:
        ws.cell(r, 1, "PDF footer stunden (sayısal çıkarım):")
        csum = ws.cell(r, col_st, value=pdf_genel_stunde)
        csum.number_format = "#,##0.0"
        csum.border = BORDER_ALL
        ws.cell(r, 1).font = DATA_FONT
        r += 1

    grand_r = r + 1
    lj = get_column_letter(col_t)
    ws.cell(grand_r, 1, "EXCEL GENEL TOPLAM (hesaplanan Toplam sütunu)").font = TITLE_FONT
    ws.merge_cells(start_row=grand_r, start_column=1, end_row=grand_r, end_column=col_t - 1)
    if employees:
        ws.cell(grand_r, col_t, value=f"=SUM({lj}2:{lj}{last_data})")
        ws.cell(grand_r, col_t).number_format = '#,##0.00 [$€-40C]'
        ws.cell(grand_r, col_t).font = Font(bold=True, size=11)

    note_r = grand_r + 2
    ws.cell(
        note_r,
        1,
        value=(
            "Stunden = SUM(günler): yalnızca sayılar toplanır (- ve metin Sİ/Ç/D dahil edilmez). "
            "Kontrol: PDF Toplam #WERT! ise PDF_HATA; sayı ise hem Stunden hem Toplam PDF ile "
            f"±{TOLERANCE} içinde olmalı (PDF Stunden sayı değilse yalnızca Toplam karşılaştırılır — formül mantığına bakın)."
        ),
    )
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=col_k)
    ws.cell(note_r, 1).font = Font(italic=True, size=8)
    ws.cell(note_r, 1).alignment = Alignment(wrap_text=True)

    last_row = note_r + 2
    apply_print_a4_landscape(ws, last_row, col_k)

    ws.freeze_panes = ws.cell(2, day_start)

    # Koşullu biçimlendirme: Kontrol sütunu
    if employees:
        col_k_letter = get_column_letter(col_k)
        rng = f"{col_k_letter}2:{col_k_letter}{last_data}"
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'${col_k_letter}2="OK"'], fill=CF_OK),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'${col_k_letter}2="HATA"'], fill=CF_HATA),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'${col_k_letter}2="PDF_HATA"'], fill=CF_PDF_HATA),
        )

    autofit_columns(ws, col_k)


def extract_pdf_footer_stunde(footer_row: list[Any] | None, cm: ColumnMap) -> float | None:
    if not footer_row or cm.col_stunde is None or cm.col_stunde >= len(footer_row):
        return None
    return parse_eu_number(str(footer_row[cm.col_stunde]))


def save_workbook(wb: Workbook) -> None:
    wb.save(OUT_PATH)
    LOG.info("Kaydedildi: %s", OUT_PATH)
    if sys.platform == "win32":
        try:
            wb.save(OUT_PATH_DOWNLOADS)
            LOG.info("Kaydedildi: %s", OUT_PATH_DOWNLOADS)
        except OSError as e:
            LOG.warning("İndirilenler kopyası yazılamadı: %s", e)


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    path = resolve_pdf_path()
    table, text_lines = extract_table_and_text(path)

    hidx = detect_header_row_index(table)
    cm = build_column_map(table[hidx])
    LOG.info("Başlık satırı indeksi=%s, gün sütun sayısı=%s", hidx, len(cm.day_cols))

    employees, footer_idx = parse_employee_rows(table, hidx, cm)
    if not employees:
        LOG.warning("Çalışan satırı bulunamadı; ham tablo yine de aktarılacak.")
    footer_pdf = parse_footer_row(table, footer_idx)
    legend = parse_legend_lines(table, footer_idx, text_lines)
    pdf_st_sum = extract_pdf_footer_stunde(footer_pdf, cm)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_pdf_sheet(wb, table, legend)
    build_working_sheet(wb, employees, footer_pdf, pdf_st_sum)

    save_workbook(wb)


if __name__ == "__main__":
    try:
        main()
    except Exception as ex:
        LOG.error("%s", ex)
        sys.exit(1)
