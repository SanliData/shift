# -*- coding: utf-8 -*-
"""
Workforce Analytics & Payroll System — Excel üretici.

Çalıştır:  python build_workforce_analytics_xlsx.py

Çıktı:     Workforce_Analytics_Payroll.xlsx (TALIP + İndirilenler)
"""
from __future__ import annotations

import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

SCRIPT_DIR = Path(__file__).resolve().parent
OUT = SCRIPT_DIR / "Workforce_Analytics_Payroll.xlsx"
OUT_DL = Path.home() / "Downloads" / "Workforce_Analytics_Payroll.xlsx"

DATA_LAST = 2000
EMP_MASTER_START = 2
EMP_MASTER_END = 52
DAILY_ROWS = 90
EMP_SUMMARY_ROWS = 35

# Örnek çalışan listesi (dropdown + özet satırları)
SAMPLE_EMPLOYEES = [
    "Ismail karakütük",
    "Burhan karakoc",
    "Murat Keles",
    "Muhittin Keles",
    "Lokman Acer",
    "Ibrahim acer",
    "Zafer kus",
    "Abdurahman ceken",
    "Murat Sakar",
    "Ümit Rugala",
    "Özdemir Hacer",
    "Yusuf islak",
    "Hakan",
    "Yusuf aygün",
    "Melih",
    "Cumali",
    "Kemal",
    "Ugur Celiker",
    "Emre",
    "Selahattin Ayyildiz",
    "hasan",
]

THIN = Side(style="thin", color="FFCCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="FF1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=18, color="FF1F4E79")
SUB_FONT = Font(size=10, color="FF666666")
KPI_LABEL = Font(bold=True, size=11)
KPI_VALUE = Font(bold=True, size=22, color="FF1F4E79")
CARD_FILL = PatternFill("solid", fgColor="FFF8F9FA")
ACCENT = PatternFill("solid", fgColor="FFE7F1FF")


def add_defined_names(wb: Workbook) -> None:
    wb.defined_names["EmployeeList"] = DefinedName(
        "EmployeeList",
        attr_text=f"'HESAPLAMA'!$Z${EMP_MASTER_START}:$Z${EMP_MASTER_END}",
    )
    wb.defined_names["VeriTarih"] = DefinedName(
        "VeriTarih", attr_text=f"'VERI_GIRISI'!$A$2:$A${DATA_LAST}"
    )
    wb.defined_names["VeriCalisan"] = DefinedName(
        "VeriCalisan", attr_text=f"'VERI_GIRISI'!$B$2:$B${DATA_LAST}"
    )
    wb.defined_names["VeriNormal"] = DefinedName(
        "VeriNormal", attr_text=f"'VERI_GIRISI'!$C$2:$C${DATA_LAST}"
    )
    wb.defined_names["VeriMesai"] = DefinedName(
        "VeriMesai", attr_text=f"'VERI_GIRISI'!$D$2:$D${DATA_LAST}"
    )
    wb.defined_names["VeriUcret"] = DefinedName(
        "VeriUcret", attr_text=f"'VERI_GIRISI'!$E$2:$E${DATA_LAST}"
    )
    wb.defined_names["VeriSatirMaliyet"] = DefinedName(
        "VeriSatirMaliyet", attr_text=f"'VERI_GIRISI'!$G$2:$G${DATA_LAST}"
    )


def build_veri_girisi(ws) -> None:
    headers = [
        "Date",
        "Employee",
        "Normal Hours",
        "Overtime Hours",
        "Hourly Rate",
        "Notes",
        "Line Cost (auto)",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(2, DATA_LAST + 1):
        ws.cell(r, 7, value=f"=IF(OR($C{r}=\"\",$E{r}=\"\"),0,$C{r}*$E{r}+$D{r}*$E{r}*1.5)")
        c = ws.cell(r, 7)
        c.fill = ACCENT
        c.font = Font(italic=True, size=9, color="FF444444")
        c.border = BORDER
        for col in range(1, 7):
            ws.cell(r, col).border = BORDER
        ws.cell(r, 1).number_format = "yyyy-mm-dd"
        ws.cell(r, 3).number_format = "0.00"
        ws.cell(r, 4).number_format = "0.00"
        ws.cell(r, 5).number_format = '#,##0.00 [$€-40C]'
        ws.cell(r, 7).number_format = '#,##0.00 [$€-40C]'

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 16

    dv_emp = DataValidation(
        type="list",
        formula1="=EmployeeList",
        allow_blank=True,
    )
    dv_emp.error = "Listeden çalışan seçin veya boş bırakın."
    dv_emp.errorTitle = "Çalışan"
    ws.add_data_validation(dv_emp)
    dv_emp.add(f"B2:B{DATA_LAST}")

    for col_letter, col_idx in [("C", 3), ("D", 4)]:
        dv_h = DataValidation(
            type="decimal",
            operator="between",
            formula1=0,
            formula2=24,
            allow_blank=True,
        )
        dv_h.error = "0 ile 24 saat arası girin."
        dv_h.errorTitle = "Saat"
        ws.add_data_validation(dv_h)
        dv_h.add(f"{col_letter}2:{col_letter}{DATA_LAST}")

    red_fill = PatternFill("solid", fgColor="FFFFE0E0")
    ws.conditional_formatting.add(
        f"C2:C{DATA_LAST}",
        CellIsRule(operator="greaterThan", formula=["10"], fill=red_fill),
    )
    ws.conditional_formatting.add(
        f"D2:D{DATA_LAST}",
        CellIsRule(operator="greaterThan", formula=["10"], fill=red_fill),
    )

    ws.freeze_panes = "A2"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"
    ws.page_margins = PageMargins(0.4, 0.4, 0.5, 0.5)


def build_hesaplama(ws) -> int:
    ws.cell(1, 1, "Çalışan özeti (SUMIFS — VERI_GIRISI)")
    ws.cell(1, 1).font = Font(bold=True, size=14, color="FF1F4E79")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    hdr = ["Employee", "Total Normal", "Total OT", "Total Hours", "Blended Rate", "Total Cost"]
    hr = 3
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(hr, c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    first = hr + 1
    last = first + EMP_SUMMARY_ROWS - 1
    for r in range(first, last + 1):
        zr = EMP_MASTER_START + (r - first)
        ws.cell(r, 1, value=f"=IF(TRIM(Z{zr})=\"\",\"\",Z{zr})")
        ws.cell(r, 2, value=f'=IF($A{r}="","",SUMIFS(VeriNormal,VeriCalisan,$A{r}))')
        ws.cell(r, 3, value=f'=IF($A{r}="","",SUMIFS(VeriMesai,VeriCalisan,$A{r}))')
        ws.cell(r, 4, value=f"=IF($A{r}=\"\",\"\",B{r}+C{r})")
        ws.cell(r, 6, value=f'=IF($A{r}="","",SUMIFS(VeriSatirMaliyet,VeriCalisan,$A{r}))')
        ws.cell(r, 5, value=f'=IF(OR($A{r}="",$D{r}=0),"",$F{r}/$D{r})')
        for c in range(1, 7):
            ws.cell(r, c).border = BORDER
        ws.cell(r, 2).number_format = "0.0"
        ws.cell(r, 3).number_format = "0.0"
        ws.cell(r, 4).number_format = "0.0"
        ws.cell(r, 5).number_format = '#,##0.00 [$€-40C]'
        ws.cell(r, 6).number_format = '#,##0.00 [$€-40C]'

    ws.cell(1, 8, "Çalışan ana listesi (dropdown kaynağı)")
    ws.cell(1, 8).font = Font(bold=True, size=10)
    ws.cell(1, 26, "Employee")
    ws.cell(1, 26).font = HDR_FONT
    ws.cell(1, 26).fill = HDR_FILL
    for i, name in enumerate(SAMPLE_EMPLOYEES):
        ws.cell(EMP_MASTER_START + i, 26, value=name)
        ws.cell(EMP_MASTER_START + i, 26).border = BORDER
    for r in range(EMP_MASTER_START + len(SAMPLE_EMPLOYEES), EMP_MASTER_END + 1):
        ws.cell(r, 26).border = BORDER

    # Günlük özet (takvim sırası — grafik beslemesi)
    dr0 = last + 4
    ws.cell(dr0, 13, "Daily rollup (for charts)")
    ws.cell(dr0, 13).font = Font(bold=True, size=12)
    ws.cell(dr0 + 1, 13, "Date")
    ws.cell(dr0 + 1, 14, "Total Hours")
    ws.cell(dr0 + 1, 15, "Total Cost")
    for c in range(13, 16):
        x = ws.cell(dr0 + 1, c)
        x.font = HDR_FONT
        x.fill = HDR_FILL
        x.border = BORDER

    start_d = date(2026, 1, 1)
    for i in range(DAILY_ROWS):
        r = dr0 + 2 + i
        d = start_d + timedelta(days=i)
        ws.cell(r, 13, value=d)
        ws.cell(r, 13).number_format = "yyyy-mm-dd"
        ws.cell(
            r,
            14,
            value=(
                f'=SUMIFS(VeriNormal,VeriTarih,M{r})+SUMIFS(VeriMesai,VeriTarih,M{r})'
            ),
        )
        ws.cell(r, 15, value=f"=SUMIFS(VeriSatirMaliyet,VeriTarih,M{r})")
        for c in range(13, 16):
            ws.cell(r, c).border = BORDER
        ws.cell(r, 14).number_format = "0.0"
        ws.cell(r, 15).number_format = '#,##0.00 [$€-40C]'

    # Aylık özet (ay başı tarihleri)
    mr0 = dr0 + DAILY_ROWS + 4
    ws.cell(mr0, 13, "Monthly totals")
    ws.cell(mr0, 13).font = Font(bold=True, size=12)
    months = [
        date(2026, 1, 1),
        date(2026, 2, 1),
        date(2026, 3, 1),
        date(2026, 4, 1),
        date(2026, 5, 1),
        date(2026, 6, 1),
    ]
    ws.cell(mr0 + 1, 13, "Month")
    ws.cell(mr0 + 1, 14, "Hours")
    ws.cell(mr0 + 1, 15, "Cost")
    for c in range(13, 16):
        ws.cell(mr0 + 1, c).font = HDR_FONT
        ws.cell(mr0 + 1, c).fill = HDR_FILL
        ws.cell(mr0 + 1, c).border = BORDER
    for i, m in enumerate(months):
        r = mr0 + 2 + i
        ws.cell(r, 13, value=m)
        ws.cell(r, 13).number_format = "mmm yyyy"
        ws.cell(
            r,
            14,
            value=(
                f'=SUMIFS(VeriNormal,VeriTarih,">="&DATE(YEAR(M{r}),MONTH(M{r}),1),'
                f'VeriTarih,"<"&DATE(YEAR(M{r}),MONTH(M{r})+1,1))'
                f'+SUMIFS(VeriMesai,VeriTarih,">="&DATE(YEAR(M{r}),MONTH(M{r}),1),'
                f'VeriTarih,"<"&DATE(YEAR(M{r}),MONTH(M{r})+1,1))'
            ),
        )
        ws.cell(
            r,
            15,
            value=(
                f'=SUMIFS(VeriSatirMaliyet,VeriTarih,">="&DATE(YEAR(M{r}),MONTH(M{r}),1),'
                f'VeriTarih,"<"&DATE(YEAR(M{r}),MONTH(M{r})+1,1))'
            ),
        )
        for c in range(13, 16):
            ws.cell(r, c).border = BORDER
        ws.cell(r, 14).number_format = "0.0"
        ws.cell(r, 15).number_format = '#,##0.00 [$€-40C]'

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["Z"].width = 22

    ws.freeze_panes = ws.cell(first, 1)
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "3:3"
    ws.page_margins = PageMargins(0.4, 0.4, 0.5, 0.5)

    return dr0 + 2


def build_dashboard(wb: Workbook, ws, hesap, chart_data_start: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, "Workforce Analytics & Payroll").font = TITLE_FONT
    ws.cell(2, 1, "Canlı özet — tüm rakamlar VERI_GIRISI üzerinden").font = SUB_FONT
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")

    g_last = DATA_LAST
    # KPI kartları
    cards = [
        ("B4", "D4", "Total Labor Cost (€)", f"=SUM('VERI_GIRISI'!G2:G{g_last})"),
        ("F4", "H4", "Total Hours Worked", f"=SUM('VERI_GIRISI'!C2:C{g_last})+SUM('VERI_GIRISI'!D2:D{g_last})"),
        ("B7", "D7", "Avg Hourly Rate (€)", f"=IFERROR((SUM('VERI_GIRISI'!G2:G{g_last}))/(SUM('VERI_GIRISI'!C2:C{g_last})+SUM('VERI_GIRISI'!D2:D{g_last})),0)"),
        ("F7", "H7", "Active Employees", "=COUNTIFS('HESAPLAMA'!$F$4:$F$38,\">0\")"),
    ]
    for a1, a2, label, formula in cards:
        r1, c1 = ws[a1].row, ws[a1].column
        r2, c2 = ws[a2].row, ws[a2].column
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(r, c)
                cell.fill = CARD_FILL
                cell.border = BORDER
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        top = ws.cell(r1, c1)
        top.value = label
        top.font = KPI_LABEL
        top.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        val_cell = ws.cell(r1 + 1, c1)
        val_cell.value = formula
        val_cell.font = KPI_VALUE
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        if "Cost" in label or "Rate" in label:
            val_cell.number_format = '#,##0.00 [$€-40C]'
        elif "Hours" in label:
            val_cell.number_format = "#,##0.0"
        else:
            val_cell.number_format = "0"

    # Top 5 çalışan (maliyet) — grafik + tablo
    ws.cell(10, 1, "Top 5 — Cost").font = Font(bold=True, size=12)
    ws.cell(10, 10, "Cost distribution (Top 5)").font = Font(bold=True, size=12)

    for i in range(5):
        r = 11 + i
        ws.cell(r, 1, value=i + 1)
        ws.cell(
            r,
            2,
            value=(
                f'=IFERROR(INDEX(\'HESAPLAMA\'!$A$4:$A$38,'
                f'MATCH(LARGE(\'HESAPLAMA\'!$F$4:$F$38,{i+1}),'
                f'\'HESAPLAMA\'!$F$4:$F$38,0)),"")'
            ),
        )
        ws.cell(
            r,
            3,
            value=f'=IFERROR(LARGE(\'HESAPLAMA\'!$F$4:$F$38,{i+1}),"")',
        )
        ws.cell(r, 3).number_format = '#,##0.00 [$€-40C]'
        for c in range(1, 4):
            ws.cell(r, c).border = BORDER

    # Önizleme tablosu (grafiklerin altında — çakışma yok)
    prev_top = 52
    ws.cell(prev_top - 1, 1, "Recent entries (preview)").font = Font(bold=True, size=12)
    prev_hdr = ["Date", "Employee", "Normal", "OT", "Rate", "Notes"]
    for c, h in enumerate(prev_hdr, 1):
        cell = ws.cell(prev_top, c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BORDER
    for r in range(10):
        src = 2 + r
        for c in range(1, 7):
            ws.cell(prev_top + 1 + r, c, value=f"='VERI_GIRISI'!{get_column_letter(c)}{src}")
            ws.cell(prev_top + 1 + r, c).border = BORDER
        ws.cell(prev_top + 1 + r, 1).number_format = "yyyy-mm-dd"
        ws.cell(prev_top + 1 + r, 3).number_format = "0.0"
        ws.cell(prev_top + 1 + r, 4).number_format = "0.0"
        ws.cell(prev_top + 1 + r, 5).number_format = '#,##0.00 [$€-40C]'

    data_start = chart_data_start
    data_end = data_start + DAILY_ROWS - 1

    cats = Reference(hesap, min_col=13, min_row=data_start, max_row=data_end)
    cost_vals = Reference(hesap, min_col=15, min_row=data_start, max_row=data_end)
    hrs_vals = Reference(hesap, min_col=14, min_row=data_start, max_row=data_end)

    line = LineChart()
    line.title = "Daily Cost Trend"
    line.y_axis.title = "Cost (€)"
    line.x_axis.title = "Date"
    line.add_data(cost_vals, titles_from_data=False)
    line.set_categories(cats)
    line.height = 9
    line.width = 18
    line.legend = None
    ws.add_chart(line, "B20")

    bar = BarChart()
    bar.type = "col"
    bar.title = "Hours vs Cost (dual axis)"
    bar.y_axis.title = "Hours"
    bar.add_data(hrs_vals, titles_from_data=False)
    bar.set_categories(cats)
    bar.height = 9
    bar.width = 18

    line2 = LineChart()
    line2.y_axis.axId = 200
    line2.y_axis.title = "Cost (€)"
    line2.add_data(cost_vals, titles_from_data=False)
    line2.y_axis.crosses = "max"
    bar.y_axis.crosses = "min"
    bar += line2
    ws.add_chart(bar, "B38")

    pie = PieChart()
    pie.title = "Top 5 Cost Share"
    labels = Reference(ws, min_col=2, min_row=11, max_row=15)
    pdata = Reference(ws, min_col=3, min_row=11, max_row=15)
    pie.add_data(pdata, titles_from_data=False)
    pie.set_categories(labels)
    pie.height = 10
    pie.width = 12
    for s in pie.series:
        s.dLbls = DataLabelList()
        s.dLbls.showPercent = True
    ws.add_chart(pie, "J11")

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 22
    for c in "CDEFGHIJKLMN":
        ws.column_dimensions[c].width = 12

    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(0.5, 0.5, 0.6, 0.6)


def main() -> None:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws_veri = wb.create_sheet("VERI_GIRISI", 0)
    ws_hesap = wb.create_sheet("HESAPLAMA", 1)
    ws_dash = wb.create_sheet("DASHBOARD", 2)

    add_defined_names(wb)
    build_veri_girisi(ws_veri)
    chart_row0 = build_hesaplama(ws_hesap)
    build_dashboard(wb, ws_dash, ws_hesap, chart_row0)

    # Özette istenen sıra: DASHBOARD → VERI_GIRISI → HESAPLAMA
    if len(wb.sheetnames) >= 3:
        wb.move_sheet(wb["DASHBOARD"], offset=-2)

    wb.save(OUT)
    print("Saved:", OUT)
    if sys.platform == "win32":
        try:
            wb.save(OUT_DL)
            print("Saved:", OUT_DL)
        except OSError as e:
            print("Downloads copy skipped:", e)


if __name__ == "__main__":
    main()
