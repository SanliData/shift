import csv
import io
from datetime import datetime
from secrets import token_urlsafe
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, File, Form, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, Response
from fastapi.templating import Jinja2Templates
from sqlalchemy import desc, func, select
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import Device, Employee, ImportedFile, RegistrationToken, TimeEntry, Vehicle

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")
BERLIN_TZ = ZoneInfo("Europe/Berlin")
DEVICE_COOKIE = "device_token"


def now_berlin() -> datetime:
    return datetime.now(BERLIN_TZ)


def parse_date(date_str: str | None):
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=BERLIN_TZ)
    except ValueError:
        return None


def build_filters(date_from: str | None, date_to: str | None, employee_id: int | None, vehicle_id: int | None):
    filters = []
    start = parse_date(date_from)
    end = parse_date(date_to)
    if start:
        filters.append(TimeEntry.start_time >= start)
    if end:
        filters.append(TimeEntry.start_time <= end.replace(hour=23, minute=59, second=59))
    if employee_id:
        filters.append(TimeEntry.employee_id == employee_id)
    if vehicle_id:
        filters.append(TimeEntry.vehicle_id == vehicle_id)
    return filters


@router.get("/admin-time", response_class=HTMLResponse)
def admin_page(
    request: Request,
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    employee_id: int | None = Query(default=None),
    vehicle_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
):
    employees = db.scalars(select(Employee).order_by(Employee.name)).all()
    vehicles = db.scalars(select(Vehicle).order_by(Vehicle.qr_code_slug)).all()
    devices = db.scalars(select(Device).order_by(desc(Device.created_at))).all()

    filters = build_filters(date_from, date_to, employee_id, vehicle_id)
    active_stmt = select(TimeEntry).where(TimeEntry.status == "active").order_by(desc(TimeEntry.start_time))
    completed_stmt = select(TimeEntry).where(TimeEntry.status == "completed").order_by(desc(TimeEntry.end_time)).limit(200)
    for f in filters:
        active_stmt = active_stmt.where(f)
        completed_stmt = completed_stmt.where(f)
    active_entries = db.scalars(active_stmt).all()
    completed_entries = db.scalars(completed_stmt).all()
    totals = db.execute(
        select(
            func.coalesce(func.sum(TimeEntry.total_minutes), 0),
            func.coalesce(func.sum(TimeEntry.overtime_minutes), 0),
        ).where(TimeEntry.status == "completed")
    ).one()
    return templates.TemplateResponse(
        "admin_time.html",
        {
            "request": request,
            "employees": employees,
            "vehicles": vehicles,
            "devices": devices,
            "active_entries": active_entries,
            "completed_entries": completed_entries,
            "total_minutes": int(totals[0] or 0),
            "overtime_minutes": int(totals[1] or 0),
            "filters": {
                "date_from": date_from or "",
                "date_to": date_to or "",
                "employee_id": employee_id or "",
                "vehicle_id": vehicle_id or "",
            },
            "register_link": "",
            "message": "",
        },
    )


@router.post("/admin-time/register-link", response_class=HTMLResponse)
def create_register_link(request: Request, employee_id: int = Form(...), db: Session = Depends(get_db)):
    employees = db.scalars(select(Employee).order_by(Employee.name)).all()
    vehicles = db.scalars(select(Vehicle).order_by(Vehicle.qr_code_slug)).all()
    devices = db.scalars(select(Device).order_by(desc(Device.created_at))).all()
    active_entries = db.scalars(select(TimeEntry).where(TimeEntry.status == "active").order_by(desc(TimeEntry.start_time))).all()
    completed_entries = db.scalars(
        select(TimeEntry).where(TimeEntry.status == "completed").order_by(desc(TimeEntry.end_time)).limit(200)
    ).all()

    token = token_urlsafe(24)
    db.add(
        RegistrationToken(
            employee_id=employee_id,
            token=token,
            used=False,
            created_at=now_berlin(),
        )
    )
    db.commit()
    link = str(request.base_url).rstrip("/") + f"/register-device?token={token}"
    return templates.TemplateResponse(
        "admin_time.html",
        {
            "request": request,
            "employees": employees,
            "vehicles": vehicles,
            "devices": devices,
            "active_entries": active_entries,
            "completed_entries": completed_entries,
            "total_minutes": sum(t.total_minutes or 0 for t in completed_entries),
            "overtime_minutes": sum(t.overtime_minutes or 0 for t in completed_entries),
            "filters": {"date_from": "", "date_to": "", "employee_id": "", "vehicle_id": ""},
            "register_link": link,
            "message": "Kayıt linki oluşturuldu.",
        },
    )


@router.get("/register-device", response_class=HTMLResponse)
def register_device(request: Request, token: str, db: Session = Depends(get_db)):
    reg = db.scalar(select(RegistrationToken).where(RegistrationToken.token == token))
    if not reg or reg.used:
        return templates.TemplateResponse(
            "register_status.html",
            {"request": request, "ok": False, "message": "Geçersiz veya kullanılmış token."},
        )
    new_token = token_urlsafe(32)
    db.add(
        Device(
            employee_id=reg.employee_id,
            device_token=new_token,
            created_at=now_berlin(),
            active=True,
        )
    )
    reg.used = True
    db.commit()
    resp = templates.TemplateResponse(
        "register_status.html",
        {"request": request, "ok": True, "message": "Cihaz başarıyla kaydedildi."},
    )
    resp.set_cookie(DEVICE_COOKIE, new_token, httponly=True, secure=False, samesite="lax", max_age=31536000)
    return resp


@router.get("/admin-time/import", response_class=HTMLResponse)
def import_page(request: Request, db: Session = Depends(get_db)):
    files = db.scalars(select(ImportedFile).order_by(desc(ImportedFile.created_at)).limit(20)).all()
    return templates.TemplateResponse(
        "import.html",
        {"request": request, "files": files, "message": "", "error": ""},
    )


@router.post("/admin-time/import", response_class=HTMLResponse)
async def import_upload(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    data = await file.read()
    imported_rows = 0
    error = ""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), data_only=True)
        if "VERI_GIRISI" in wb.sheetnames:
            ws = wb["VERI_GIRISI"]
            imported_rows = max(0, ws.max_row - 1)
        else:
            imported_rows = 0
        db.add(ImportedFile(filename=file.filename or "uploaded.xlsx", imported_rows=imported_rows, created_at=now_berlin()))
        db.commit()
    except Exception as ex:
        error = str(ex)
    files = db.scalars(select(ImportedFile).order_by(desc(ImportedFile.created_at)).limit(20)).all()
    return templates.TemplateResponse(
        "import.html",
        {"request": request, "files": files, "message": f"Import tamamlandı. Satır: {imported_rows}", "error": error},
    )


@router.get("/admin-time/export")
def export_xlsx(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    entries = db.scalars(select(TimeEntry).order_by(desc(TimeEntry.start_time))).all()
    employees = db.scalars(select(Employee).order_by(Employee.id)).all()
    devices = db.scalars(select(Device)).all()

    wb = Workbook()
    sheets = {
        "DASHBOARD": wb.active,
        "PERSONEL": wb.create_sheet("PERSONEL"),
        "VERI_GIRISI": wb.create_sheet("VERI_GIRISI"),
        "HESAPLAMA": wb.create_sheet("HESAPLAMA"),
    }
    sheets["DASHBOARD"].title = "DASHBOARD"

    fill = PatternFill("solid", fgColor="1F4E79")
    white = Font(color="FFFFFF", bold=True)

    total_minutes = sum(e.total_minutes or 0 for e in entries)
    total_overtime = sum(e.overtime_minutes or 0 for e in entries)
    active_count = sum(1 for e in entries if e.status == "active")
    ws = sheets["DASHBOARD"]
    ws.append(["KPI", "Değer"])
    ws.append(["Toplam Çalışan", len(employees)])
    ws.append(["Toplam Saat", round(total_minutes / 60, 2)])
    ws.append(["Fazla Mesai Saat", round(total_overtime / 60, 2)])
    ws.append(["Aktif Mesai", active_count])

    for name, ws in sheets.items():
        if name == "DASHBOARD":
            continue
        if name == "PERSONEL":
            ws.append(["employee_id", "employee_name", "active", "registered_device_count", "total_hours", "overtime_hours"])
            for emp in employees:
                device_count = sum(1 for d in devices if d.employee_id == emp.id and d.active)
                emp_entries = [e for e in entries if e.employee_id == emp.id]
                ws.append(
                    [
                        emp.id,
                        emp.name,
                        "yes" if emp.active else "no",
                        device_count,
                        round(sum((e.total_minutes or 0) for e in emp_entries) / 60, 2),
                        round(sum((e.overtime_minutes or 0) for e in emp_entries) / 60, 2),
                    ]
                )
        elif name == "VERI_GIRISI":
            ws.append(["tarih", "çalışan", "araç", "başlangıç", "bitiş", "toplam dakika", "fazla mesai", "durum"])
            for e in entries:
                ws.append(
                    [
                        str(e.start_time.date()) if e.start_time else "",
                        e.employee_name,
                        e.vehicle_id,
                        str(e.start_time or ""),
                        str(e.end_time or ""),
                        e.total_minutes or 0,
                        e.overtime_minutes or 0,
                        e.status,
                    ]
                )
        elif name == "HESAPLAMA":
            ws.append(["employee_name", "total_hours", "overtime_hours"])
            for emp in employees:
                emp_entries = [e for e in entries if e.employee_id == emp.id]
                ws.append(
                    [
                        emp.name,
                        round(sum((e.total_minutes or 0) for e in emp_entries) / 60, 2),
                        round(sum((e.overtime_minutes or 0) for e in emp_entries) / 60, 2),
                    ]
                )

    for ws in sheets.values():
        ws.auto_filter.ref = f"A1:{chr(64 + ws.max_column)}1"
        for cell in ws[1]:
            cell.fill = fill
            cell.font = white

    out = io.BytesIO()
    wb.save(out)
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=qr_time_export.xlsx"},
    )
