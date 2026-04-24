import base64
import csv
import io
import json
import os
from datetime import datetime, timedelta
from secrets import token_urlsafe
from zoneinfo import ZoneInfo

import qrcode
from fastapi import Depends, FastAPI, Form, Query, Request
from fastapi import File, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import Integer, desc, func, select, text
from sqlalchemy.orm import Session, joinedload

from .db import Base, SessionLocal, engine, get_db
from .models import (
    Device,
    Employee,
    ImportedFile,
    MonthlySummary,
    PayrollExport,
    RegistrationToken,
    TimeEntry,
    Vehicle,
)

BERLIN_TZ = ZoneInfo("Europe/Berlin")
DEVICE_COOKIE = "device_token"
EXTERNAL_REDIRECT_URL = "https://novarchive.org/ui/index.html"
COOKIE_SECURE = os.getenv("COOKIE_SECURE", "false").strip().lower() in ("1", "true", "yes")

app = FastAPI(title="Sanli Netzbau Time Tracker")
app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")


def now_berlin() -> datetime:
    return datetime.now(BERLIN_TZ)


def as_berlin(dt: datetime) -> datetime:
    if dt.tzinfo is None:
        return dt.replace(tzinfo=BERLIN_TZ)
    return dt.astimezone(BERLIN_TZ)


def get_registered_device(db: Session, request: Request) -> Device | None:
    device_token = request.cookies.get(DEVICE_COOKIE)
    if not device_token:
        return None
    stmt = (
        select(Device)
        .options(joinedload(Device.employee))
        .where(Device.device_token == device_token, Device.active.is_(True))
    )
    device = db.scalar(stmt)
    if not device or not device.employee or not device.employee.active:
        return None
    return device


def get_active_entry(db: Session, employee_id: int) -> TimeEntry | None:
    stmt = (
        select(TimeEntry)
        .where(
            TimeEntry.employee_id == employee_id,
            TimeEntry.status == "active",
        )
        .order_by(desc(TimeEntry.start_time))
    )
    return db.scalar(stmt)


def parse_date(date_str: str | None) -> datetime | None:
    if not date_str:
        return None
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return dt.replace(tzinfo=BERLIN_TZ)
    except ValueError:
        return None


def build_admin_filtered_query(
    date_from: str | None,
    date_to: str | None,
    employee_id: int | None,
    vehicle_id: int | None,
):
    filters = []
    dt_from = parse_date(date_from)
    dt_to = parse_date(date_to)
    if dt_from:
        filters.append(TimeEntry.start_time >= dt_from)
    if dt_to:
        filters.append(TimeEntry.start_time < dt_to.replace(hour=23, minute=59, second=59))
    if employee_id:
        filters.append(TimeEntry.employee_id == employee_id)
    if vehicle_id:
        filters.append(TimeEntry.vehicle_id == vehicle_id)
    return filters


def build_report_filters(
    date_from: str | None,
    date_to: str | None,
    employee_id: int | None,
    vehicle_id: int | None,
    month: int | None,
    status: str | None,
):
    filters = build_admin_filtered_query(date_from, date_to, employee_id, vehicle_id)
    if month and 1 <= month <= 12:
        filters.append(func.cast(func.strftime("%m", TimeEntry.start_time), Integer) == month)
    if status and status in ("active", "completed"):
        filters.append(TimeEntry.status == status)
    return filters


def refresh_monthly_summaries(db: Session) -> None:
    db.query(MonthlySummary).delete()
    entries = db.scalars(select(TimeEntry)).all()
    grouped: dict[tuple[int, int], dict[str, int]] = {}
    for e in entries:
        if not e.start_time:
            continue
        dt = as_berlin(e.start_time)
        key = (dt.year, dt.month)
        grouped.setdefault(
            key,
            {
                "total_minutes": 0,
                "overtime_minutes": 0,
                "active_entries": 0,
                "completed_entries": 0,
                "missing_checkout_entries": 0,
            },
        )
        grouped[key]["total_minutes"] += e.total_minutes or 0
        grouped[key]["overtime_minutes"] += e.overtime_minutes or 0
        if e.status == "active":
            grouped[key]["active_entries"] += 1
            grouped[key]["missing_checkout_entries"] += 1
        elif e.status == "completed":
            grouped[key]["completed_entries"] += 1
            if not e.end_time:
                grouped[key]["missing_checkout_entries"] += 1

    now = now_berlin()
    for (year, month), values in grouped.items():
        db.add(
            MonthlySummary(
                year=year,
                month=month,
                employee_id=None,
                vehicle_id=None,
                total_minutes=values["total_minutes"],
                overtime_minutes=values["overtime_minutes"],
                active_entries=values["active_entries"],
                completed_entries=values["completed_entries"],
                missing_checkout_entries=values["missing_checkout_entries"],
                created_at=now,
            )
        )
    db.commit()


def normalize_header(value: str) -> str:
    return value.strip().lower().replace("_", " ")


def import_personel_excel(db: Session, file_content: bytes, filename: str) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    if "VERI_GIRISI" not in wb.sheetnames:
        return 0
    ws = wb["VERI_GIRISI"]
    header_cells = [normalize_header(str(c.value or "")) for c in ws[1]]
    idx = {h: i for i, h in enumerate(header_cells)}

    # esnek başlık eşleme
    def col(*candidates):
        for c in candidates:
            if c in idx:
                return idx[c]
        return None

    date_i = col("date", "tarih")
    emp_i = col("employee", "çalışan", "calisan", "ad soyad")
    normal_i = col("normal hours", "normal saat", "normal")
    over_i = col("overtime hours", "fazla mesai", "mesai")
    row_cost_i = col("line cost (auto)", "line cost")

    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        emp_name = str(row[emp_i]).strip() if emp_i is not None and row[emp_i] else ""
        if not emp_name:
            continue
        employee = db.scalar(select(Employee).where(Employee.name == emp_name))
        if not employee:
            employee = Employee(name=emp_name, active=True)
            db.add(employee)
            db.flush()

        # demo importte device yoksa sanal device yarat
        device = db.scalar(select(Device).where(Device.employee_id == employee.id, Device.active.is_(True)))
        if not device:
            device = Device(
                employee_id=employee.id,
                device_token=f"imported-{employee.id}-{token_urlsafe(8)}",
                created_at=now_berlin(),
                active=True,
            )
            db.add(device)
            db.flush()

        vehicle = db.scalar(select(Vehicle).where(Vehicle.qr_code_slug == "vehicle-01"))
        if not vehicle:
            vehicle = Vehicle(name="vehicle-01", qr_code_slug="vehicle-01")
            db.add(vehicle)
            db.flush()

        raw_date = row[date_i] if date_i is not None else None
        if isinstance(raw_date, datetime):
            start_dt = raw_date.replace(tzinfo=BERLIN_TZ)
        else:
            parsed = parse_date(str(raw_date)[:10]) if raw_date else now_berlin()
            start_dt = parsed or now_berlin()
        normal = float(row[normal_i] or 0) if normal_i is not None else 0.0
        overtime = float(row[over_i] or 0) if over_i is not None else 0.0
        total_min = int((normal + overtime) * 60)
        overtime_min = int(max(0, overtime * 60))
        end_dt = start_dt + timedelta(minutes=total_min)

        db.add(
            TimeEntry(
                employee_id=employee.id,
                employee_name=employee.name,
                device_id=device.id,
                vehicle_id=vehicle.id,
                start_time=start_dt,
                end_time=end_dt,
                total_minutes=total_min,
                overtime_minutes=overtime_min,
                status="completed",
            )
        )
        imported += 1
    db.add(
        ImportedFile(
            filename=filename,
            source_type="personel_analytics",
            imported_rows=imported,
            created_at=now_berlin(),
        )
    )
    db.commit()
    refresh_monthly_summaries(db)
    return imported


def import_altyapi_excel(db: Session, file_content: bytes, filename: str) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    # Bu dosya operasyonel/aylık mantık referansı: satır importu yerine metadata tutuyoruz.
    recognized_month_sheets = [s for s in wb.sheetnames if s in {
        "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
        "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"
    }]
    imported = len(recognized_month_sheets)
    db.add(
        ImportedFile(
            filename=filename,
            source_type="altyapi_yonetim",
            imported_rows=imported,
            created_at=now_berlin(),
        )
    )
    db.commit()
    refresh_monthly_summaries(db)
    return imported


def seed_demo_data(db: Session) -> None:
    if db.scalar(select(func.count(Employee.id))) == 0:
        db.add_all(
            [
                Employee(name="Mehmet Yilmaz", active=True),
                Employee(name="Ali Demir", active=True),
            ]
        )
    if db.scalar(select(func.count(Vehicle.id))) == 0:
        db.add_all(
            [
                Vehicle(name="vehicle-01", qr_code_slug="vehicle-01"),
                Vehicle(name="vehicle-02", qr_code_slug="vehicle-02"),
            ]
        )
    db.commit()


@app.on_event("startup")
def startup() -> None:
    Base.metadata.create_all(bind=engine)
    with SessionLocal() as db:
        # Hafif migration: eski DB'de employee_name kolonu yoksa ekle.
        cols = db.execute(text("PRAGMA table_info(time_entries)")).fetchall()
        col_names = {c[1] for c in cols}
        if "employee_name" not in col_names and cols:
            db.execute(text("ALTER TABLE time_entries ADD COLUMN employee_name VARCHAR(120)"))
            db.execute(
                text(
                    "UPDATE time_entries SET employee_name = ("
                    "SELECT name FROM employees WHERE employees.id = time_entries.employee_id)"
                )
            )
            db.commit()
        seed_demo_data(db)


@app.get("/", response_class=HTMLResponse)
def home(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        "home.html",
        {
            "request": request,
            "title": "Sanli Netzbau Mesai Sistemi",
        },
    )


@app.get("/time", response_class=HTMLResponse)
def time_page(request: Request, vehicle: str, db: Session = Depends(get_db)) -> HTMLResponse:
    vehicle_obj = db.scalar(select(Vehicle).where(Vehicle.qr_code_slug == vehicle))
    if not vehicle_obj:
        return RedirectResponse(url=EXTERNAL_REDIRECT_URL, status_code=302)

    device = get_registered_device(db, request)
    if not device:
        return RedirectResponse(url=EXTERNAL_REDIRECT_URL, status_code=302)

    active_entry = get_active_entry(db, device.employee_id)
    return templates.TemplateResponse(
        "time.html",
        {
            "request": request,
            "employee": device.employee,
            "vehicle": vehicle_obj,
            "active_entry": active_entry,
            "message": request.query_params.get("message", ""),
            "error": request.query_params.get("error", ""),
            "now_time": now_berlin().strftime("%d.%m.%Y %H:%M"),
        },
    )


@app.post("/time/start")
def start_shift(
    request: Request,
    vehicle_slug: str = Form(...),
    db: Session = Depends(get_db),
):
    device = get_registered_device(db, request)
    if not device:
        return RedirectResponse(url=EXTERNAL_REDIRECT_URL, status_code=302)

    vehicle = db.scalar(select(Vehicle).where(Vehicle.qr_code_slug == vehicle_slug))
    if not vehicle:
        return RedirectResponse(url=EXTERNAL_REDIRECT_URL, status_code=302)

    if get_active_entry(db, device.employee_id):
        return RedirectResponse(
            url=f"/time?vehicle={vehicle_slug}&error=Aktif mesai zaten mevcut.",
            status_code=303,
        )

    entry = TimeEntry(
        employee_id=device.employee_id,
        employee_name=device.employee.name,
        device_id=device.id,
        vehicle_id=vehicle.id,
        start_time=now_berlin(),
        status="active",
    )
    db.add(entry)
    db.commit()
    return RedirectResponse(
        url=f"/time?vehicle={vehicle_slug}&message=Mesai başarıyla başlatıldı.",
        status_code=303,
    )


@app.post("/time/stop")
def stop_shift(
    request: Request,
    vehicle_slug: str = Form(...),
    db: Session = Depends(get_db),
):
    device = get_registered_device(db, request)
    if not device:
        return RedirectResponse(url=EXTERNAL_REDIRECT_URL, status_code=302)

    vehicle = db.scalar(select(Vehicle).where(Vehicle.qr_code_slug == vehicle_slug))
    if not vehicle:
        return RedirectResponse(url=EXTERNAL_REDIRECT_URL, status_code=302)

    active_entry = get_active_entry(db, device.employee_id)
    if not active_entry:
        return RedirectResponse(
            url=f"/time?vehicle={vehicle_slug}&error=Aktif mesai bulunamadı.",
            status_code=303,
        )

    end_time = now_berlin()
    minutes = max(0, int((end_time - as_berlin(active_entry.start_time)).total_seconds() // 60))
    active_entry.end_time = end_time
    active_entry.total_minutes = minutes
    active_entry.status = "completed"

    local_date = as_berlin(active_entry.start_time).date()
    completed_stmt = select(TimeEntry).where(
        TimeEntry.employee_id == device.employee_id,
        TimeEntry.status == "completed",
    )
    existing_entries = db.scalars(completed_stmt).all()
    existing_today_minutes = sum(
        e.total_minutes or 0
        for e in existing_entries
        if e.id != active_entry.id
        and e.start_time
        and as_berlin(e.start_time).date() == local_date
    )
    cumulative_today = existing_today_minutes + minutes
    active_entry.overtime_minutes = max(0, cumulative_today - 480)

    db.commit()
    return RedirectResponse(
        url=f"/time?vehicle={vehicle_slug}&message=Mesai başarıyla bitirildi.",
        status_code=303,
    )


@app.get("/admin", response_class=HTMLResponse)
def admin_dashboard(
    request: Request,
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    employee_id: int | None = Query(default=None),
    vehicle_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
) -> HTMLResponse:
    employees = db.scalars(select(Employee).order_by(Employee.name)).all()
    vehicles = db.scalars(select(Vehicle).order_by(Vehicle.name)).all()
    filters = build_admin_filtered_query(date_from, date_to, employee_id, vehicle_id)

    active_stmt = (
        select(TimeEntry)
        .options(joinedload(TimeEntry.employee), joinedload(TimeEntry.vehicle))
        .where(TimeEntry.status == "active")
        .order_by(desc(TimeEntry.start_time))
    )
    completed_stmt = (
        select(TimeEntry)
        .options(joinedload(TimeEntry.employee), joinedload(TimeEntry.vehicle))
        .where(TimeEntry.status == "completed")
        .order_by(desc(TimeEntry.end_time))
        .limit(100)
    )
    for f in filters:
        active_stmt = active_stmt.where(f)
        completed_stmt = completed_stmt.where(f)
    active_entries = db.scalars(active_stmt).all()
    completed_entries = db.scalars(completed_stmt).all()

    totals_stmt = select(
        func.coalesce(func.sum(TimeEntry.total_minutes), 0),
        func.coalesce(func.sum(TimeEntry.overtime_minutes), 0),
    ).where(TimeEntry.status == "completed")
    for f in filters:
        totals_stmt = totals_stmt.where(f)
    total_minutes, overtime_minutes = db.execute(totals_stmt).one()

    return templates.TemplateResponse(
        "admin.html",
        {
            "request": request,
            "employees": employees,
            "vehicles": vehicles,
            "active_entries": active_entries,
            "completed_entries": completed_entries,
            "total_minutes": int(total_minutes or 0),
            "overtime_minutes": int(overtime_minutes or 0),
            "filters": {
                "date_from": date_from or "",
                "date_to": date_to or "",
                "employee_id": employee_id or "",
                "vehicle_id": vehicle_id or "",
            },
        },
    )


@app.get("/admin-time", response_class=HTMLResponse)
def admin_time_alias(
    request: Request,
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    employee_id: int | None = Query(default=None),
    vehicle_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
) -> HTMLResponse:
    return admin_dashboard(request, date_from, date_to, employee_id, vehicle_id, db)


@app.get("/admin-time/import", response_class=HTMLResponse)
def admin_time_import_page(request: Request, db: Session = Depends(get_db)) -> HTMLResponse:
    recent_imports = db.scalars(select(ImportedFile).order_by(desc(ImportedFile.created_at)).limit(20)).all()
    return templates.TemplateResponse(
        "admin_time_import.html",
        {"request": request, "recent_imports": recent_imports, "message": "", "error": ""},
    )


@app.post("/admin-time/import", response_class=HTMLResponse)
async def admin_time_import_upload(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
) -> HTMLResponse:
    content = await file.read()
    filename = file.filename or "uploaded.xlsx"
    msg = ""
    err = ""
    try:
        if "personel_analytics_production_ready" in filename.lower():
            imported = import_personel_excel(db, content, filename)
            msg = f"{imported} mesai satırı içeri alındı (personel analytics)."
        elif "altyapi_yonetim_sistemi" in filename.lower():
            imported = import_altyapi_excel(db, content, filename)
            msg = f"{imported} aylık sayfa mantığı algılandı (altyapı yönetim)."
        else:
            # İsim farklıysa önce personel dosyası gibi dene
            imported = import_personel_excel(db, content, filename)
            msg = f"{imported} satır içeri alındı."
    except Exception as ex:
        err = f"Import sırasında hata oluştu: {ex}"
    recent_imports = db.scalars(select(ImportedFile).order_by(desc(ImportedFile.created_at)).limit(20)).all()
    return templates.TemplateResponse(
        "admin_time_import.html",
        {"request": request, "recent_imports": recent_imports, "message": msg, "error": err},
    )


@app.get("/admin-time/dashboard", response_class=HTMLResponse)
def admin_time_dashboard(request: Request, db: Session = Depends(get_db)) -> HTMLResponse:
    now = now_berlin()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    employee_count = db.scalar(select(func.count(Employee.id)).where(Employee.active.is_(True))) or 0
    active_today = db.scalar(
        select(func.count(TimeEntry.id)).where(
            TimeEntry.status == "active",
            TimeEntry.start_time >= now.replace(hour=0, minute=0, second=0, microsecond=0),
        )
    ) or 0
    month_totals = db.execute(
        select(
            func.coalesce(func.sum(TimeEntry.total_minutes), 0),
            func.coalesce(func.sum(TimeEntry.overtime_minutes), 0),
        ).where(TimeEntry.start_time >= month_start)
    ).one()
    month_minutes, month_overtime = int(month_totals[0] or 0), int(month_totals[1] or 0)

    vehicle_rows = db.execute(
        select(Vehicle.qr_code_slug, func.coalesce(func.sum(TimeEntry.total_minutes), 0))
        .join(TimeEntry, TimeEntry.vehicle_id == Vehicle.id, isouter=True)
        .group_by(Vehicle.qr_code_slug)
        .order_by(Vehicle.qr_code_slug)
    ).all()
    employee_rows = db.execute(
        select(Employee.name, func.coalesce(func.sum(TimeEntry.total_minutes), 0))
        .join(TimeEntry, TimeEntry.employee_id == Employee.id, isouter=True)
        .group_by(Employee.name)
        .order_by(Employee.name)
    ).all()
    monthly = db.scalars(select(MonthlySummary).order_by(MonthlySummary.year, MonthlySummary.month)).all()

    return templates.TemplateResponse(
        "admin_time_dashboard.html",
        {
            "request": request,
            "employee_count": employee_count,
            "active_today": active_today,
            "month_hours": round(month_minutes / 60, 2),
            "month_overtime_hours": round(month_overtime / 60, 2),
            "vehicle_rows": vehicle_rows,
            "employee_rows": employee_rows,
            "monthly_rows": monthly,
        },
    )


@app.get("/admin-time/reports", response_class=HTMLResponse)
def admin_time_reports(
    request: Request,
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    employee_id: int | None = Query(default=None),
    vehicle_id: int | None = Query(default=None),
    month: int | None = Query(default=None),
    status: str | None = Query(default=None),
    db: Session = Depends(get_db),
) -> HTMLResponse:
    filters = build_report_filters(date_from, date_to, employee_id, vehicle_id, month, status)
    stmt = (
        select(TimeEntry)
        .options(joinedload(TimeEntry.employee), joinedload(TimeEntry.vehicle))
        .order_by(desc(TimeEntry.start_time))
        .limit(300)
    )
    for f in filters:
        stmt = stmt.where(f)
    entries = db.scalars(stmt).all()
    employees = db.scalars(select(Employee).order_by(Employee.name)).all()
    vehicles = db.scalars(select(Vehicle).order_by(Vehicle.qr_code_slug)).all()
    missing_checkout = [e for e in entries if e.status == "active" or not e.end_time]
    return templates.TemplateResponse(
        "admin_time_reports.html",
        {
            "request": request,
            "entries": entries,
            "employees": employees,
            "vehicles": vehicles,
            "missing_checkout": missing_checkout,
            "filters": {
                "date_from": date_from or "",
                "date_to": date_to or "",
                "employee_id": employee_id or "",
                "vehicle_id": vehicle_id or "",
                "month": month or "",
                "status": status or "",
            },
        },
    )


@app.get("/admin/export")
def admin_export(
    format: str = Query(default="csv", pattern="^(csv|xlsx)$"),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    employee_id: int | None = Query(default=None),
    vehicle_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
):
    filters = build_admin_filtered_query(date_from, date_to, employee_id, vehicle_id)
    stmt = (
        select(TimeEntry)
        .options(joinedload(TimeEntry.employee), joinedload(TimeEntry.vehicle), joinedload(TimeEntry.device))
        .order_by(desc(TimeEntry.start_time))
    )
    for f in filters:
        stmt = stmt.where(f)
    entries = db.scalars(stmt).all()

    rows = [
        [
            "entry_id",
            "employee_id",
            "employee_name",
            "device_id",
            "vehicle",
            "start_time",
            "end_time",
            "total_minutes",
            "overtime_minutes",
            "status",
        ]
    ]
    for e in entries:
        rows.append(
            [
                e.id,
                e.employee_id,
                e.employee_name,
                e.device_id,
                e.vehicle.qr_code_slug if e.vehicle else "",
                str(e.start_time or ""),
                str(e.end_time or ""),
                e.total_minutes or 0,
                e.overtime_minutes or 0,
                e.status,
            ]
        )

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(rows)
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=time_entries_export.csv"},
        )

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_dashboard = wb.active
    ws_dashboard.title = "DASHBOARD"
    ws_personel = wb.create_sheet("PERSONEL")
    ws_veri = wb.create_sheet("VERI_GIRISI")
    ws_hesap = wb.create_sheet("HESAPLAMA")
    ws_gelir = wb.create_sheet("Gelir-Gider")
    month_names_tr = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
    month_sheets = {i + 1: wb.create_sheet(month_names_tr[i]) for i in range(12)}

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    kpi_fill = PatternFill("solid", fgColor="D9E1F2")

    def style_header(ws, row=1):
        max_col = ws.max_column
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.auto_filter.ref = f"A{row}:{get_column_letter(max_col)}{row}"

    def autofit(ws):
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            width = 10
            for row in range(1, min(ws.max_row, 600) + 1):
                val = ws.cell(row=row, column=col).value
                if val is None:
                    continue
                width = max(width, min(45, len(str(val)) + 2))
            ws.column_dimensions[letter].width = width

    total_minutes = sum(int(r[7] or 0) for r in rows[1:]) if len(rows) > 1 else 0
    total_overtime = sum(int(r[8] or 0) for r in rows[1:]) if len(rows) > 1 else 0
    active_count = sum(1 for r in rows[1:] if r[9] == "active")

    emp_agg: dict[int, dict[str, float]] = {}
    vehicle_agg: dict[str, int] = {}
    monthly_agg: dict[int, int] = {m: 0 for m in range(1, 13)}
    month_rows: dict[int, list[list]] = {m: [] for m in range(1, 13)}

    all_employees = db.scalars(select(Employee).order_by(Employee.id)).all()
    device_count_rows = db.execute(
        select(Device.employee_id, func.count(Device.id)).where(Device.active.is_(True)).group_by(Device.employee_id)
    ).all()
    device_counts = {eid: cnt for eid, cnt in device_count_rows}

    veri_rows = [
        ["tarih", "employee_id", "çalışan", "araç", "başlangıç", "bitiş", "toplam dakika", "fazla mesai", "durum"]
    ]
    for e in entries:
        mins = int(e.total_minutes or 0)
        over = int(e.overtime_minutes or 0)
        vslug = e.vehicle.qr_code_slug if e.vehicle else "unknown"
        employee_key = e.employee_id
        emp_agg.setdefault(employee_key, {"name": e.employee_name, "hours": 0.0, "overtime": 0.0})
        emp_agg[employee_key]["hours"] += mins / 60
        emp_agg[employee_key]["overtime"] += over / 60
        vehicle_agg[vslug] = vehicle_agg.get(vslug, 0) + mins

        month_num = 0
        day_val = ""
        start_val = ""
        end_val = ""
        if e.start_time:
            dt = as_berlin(e.start_time)
            month_num = dt.month
            day_val = dt.date().isoformat()
            start_val = dt.strftime("%Y-%m-%d %H:%M")
            monthly_agg[month_num] += mins
        if e.end_time:
            end_val = as_berlin(e.end_time).strftime("%Y-%m-%d %H:%M")
        row = [day_val, e.employee_id, e.employee_name, vslug, start_val, end_val, mins, over, e.status]
        veri_rows.append(row)
        if month_num in month_rows:
            month_rows[month_num].append(row)

    # DASHBOARD
    ws_dashboard.append(["KPI", "Değer"])
    ws_dashboard.append(["Toplam Çalışan", len(all_employees)])
    ws_dashboard.append(["Toplam Saat", round(total_minutes / 60, 2)])
    ws_dashboard.append(["Fazla Mesai Saat", round(total_overtime / 60, 2)])
    ws_dashboard.append(["Aktif Mesai", active_count])
    ws_dashboard.append([])
    ws_dashboard.append(["Araç", "Toplam Saat"])
    for vehicle, mins in sorted(vehicle_agg.items()):
        ws_dashboard.append([vehicle, round(mins / 60, 2)])
    style_header(ws_dashboard, 1)
    if ws_dashboard.max_row >= 7:
        style_header(ws_dashboard, 7)
    for r in range(2, 6):
        ws_dashboard.cell(r, 1).fill = kpi_fill
        ws_dashboard.cell(r, 2).fill = kpi_fill

    # PERSONEL
    ws_personel.append(["employee_id", "employee_name", "active", "registered_device_count", "total_hours", "overtime_hours"])
    for emp in all_employees:
        agg = emp_agg.get(emp.id, {"hours": 0.0, "overtime": 0.0, "name": emp.name})
        ws_personel.append(
            [
                emp.id,
                emp.name,
                "yes" if emp.active else "no",
                int(device_counts.get(emp.id, 0)),
                round(float(agg["hours"]), 2),
                round(float(agg["overtime"]), 2),
            ]
        )
    style_header(ws_personel, 1)

    # VERI_GIRISI
    for row in veri_rows:
        ws_veri.append(row)
    style_header(ws_veri, 1)

    # HESAPLAMA
    ws_hesap.append(["Çalışan Bazlı", "", ""])
    ws_hesap.append(["employee_name", "total_hours", "overtime_hours"])
    for _, vals in sorted(emp_agg.items(), key=lambda x: x[1]["name"]):
        ws_hesap.append([vals["name"], round(vals["hours"], 2), round(vals["overtime"], 2)])
    ws_hesap.append([])
    ws_hesap.append(["Aylık", "", ""])
    ws_hesap.append(["ay", "total_hours", "overtime_hours"])
    for m in range(1, 13):
        month_total_minutes = monthly_agg.get(m, 0)
        month_over_minutes = sum(int(r[7]) for r in month_rows.get(m, []))
        ws_hesap.append([month_names_tr[m - 1], round(month_total_minutes / 60, 2), round(month_over_minutes / 60, 2)])
    style_header(ws_hesap, 2)
    style_header(ws_hesap, 8)

    # Aylık sayfalar
    for month_num, sheet in month_sheets.items():
        sheet.append(["tarih", "employee_id", "çalışan", "araç", "başlangıç", "bitiş", "toplam dakika", "fazla mesai", "durum"])
        month_data = month_rows.get(month_num, [])
        if month_data:
            for row in month_data:
                sheet.append(row)
        else:
            sheet.append(["", "", "", "", "", "", 0, 0, "no_data"])
        style_header(sheet, 1)

    # Gelir-Gider
    ws_gelir.append(["Kalem", "Değer"])
    ws_gelir.append(["Saatlik Ücret (EUR)", 20.0])
    ws_gelir.append(["Fazla Mesai Çarpanı", 1.5])
    ws_gelir.append(["Toplam Çalışma Saati", round(total_minutes / 60, 2)])
    ws_gelir.append(["Toplam Fazla Mesai Saati", round(total_overtime / 60, 2)])
    ws_gelir.append(["Normal İşçilik Maliyeti", "=B2*B4"])
    ws_gelir.append(["Fazla Mesai Maliyeti", "=B2*B3*B5"])
    ws_gelir.append(["Toplam İşçilik Maliyeti", "=B6+B7"])
    style_header(ws_gelir, 1)
    ws_gelir["B2"].number_format = "#,##0.00"
    ws_gelir["B3"].number_format = "#,##0.00"
    ws_gelir["B6"].number_format = "#,##0.00"
    ws_gelir["B7"].number_format = "#,##0.00"
    ws_gelir["B8"].number_format = "#,##0.00"

    # Tarih/saat ve genişlik
    for ws in [ws_veri, *month_sheets.values()]:
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value and ws.cell(r, 1).value != "no_data":
                ws.cell(r, 1).number_format = "yyyy-mm-dd"
        autofit(ws)
    for ws in [ws_dashboard, ws_personel, ws_hesap, ws_gelir]:
        autofit(ws)

    db.add(
        PayrollExport(
            report_type="xlsx_export",
            filters_json=json.dumps(
                {
                    "date_from": date_from,
                    "date_to": date_to,
                    "employee_id": employee_id,
                    "vehicle_id": vehicle_id,
                }
            ),
            row_count=max(0, len(rows) - 1),
            created_at=now_berlin(),
        )
    )
    db.commit()
    bio = io.BytesIO()
    wb.save(bio)
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=time_entries_export.xlsx"},
    )


@app.get("/admin-time/export")
def admin_time_export(
    report: str = Query(default="daily", pattern="^(daily|weekly|monthly|employee_overtime|vehicle_site)$"),
    format: str = Query(default="xlsx", pattern="^(csv|xlsx)$"),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    employee_id: int | None = Query(default=None),
    vehicle_id: int | None = Query(default=None),
    month: int | None = Query(default=None),
    status: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    filters = build_report_filters(date_from, date_to, employee_id, vehicle_id, month, status)
    stmt = (
        select(TimeEntry)
        .options(joinedload(TimeEntry.employee), joinedload(TimeEntry.vehicle))
        .order_by(TimeEntry.start_time)
    )
    for f in filters:
        stmt = stmt.where(f)
    entries = db.scalars(stmt).all()

    if report == "employee_overtime":
        rows = [["employee", "total_hours", "overtime_hours"]]
        agg = {}
        for e in entries:
            agg.setdefault(e.employee_name, [0, 0])
            agg[e.employee_name][0] += e.total_minutes or 0
            agg[e.employee_name][1] += e.overtime_minutes or 0
        for k, v in agg.items():
            rows.append([k, round(v[0] / 60, 2), round(v[1] / 60, 2)])
    elif report == "vehicle_site":
        rows = [["vehicle", "total_hours", "entries"]]
        agg = {}
        for e in entries:
            vname = e.vehicle.qr_code_slug if e.vehicle else "unknown"
            agg.setdefault(vname, [0, 0])
            agg[vname][0] += e.total_minutes or 0
            agg[vname][1] += 1
        for k, v in agg.items():
            rows.append([k, round(v[0] / 60, 2), v[1]])
    else:
        rows = [["start_time", "employee", "vehicle", "status", "minutes", "overtime_minutes"]]
        for e in entries:
            rows.append(
                [
                    str(e.start_time or ""),
                    e.employee_name,
                    e.vehicle.qr_code_slug if e.vehicle else "",
                    e.status,
                    e.total_minutes or 0,
                    e.overtime_minutes or 0,
                ]
            )

    db.add(
        PayrollExport(
            report_type=f"{report}_{format}",
            filters_json=json.dumps(
                {
                    "date_from": date_from,
                    "date_to": date_to,
                    "employee_id": employee_id,
                    "vehicle_id": vehicle_id,
                    "month": month,
                    "status": status,
                }
            ),
            row_count=max(0, len(rows) - 1),
            created_at=now_berlin(),
        )
    )
    db.commit()

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(rows)
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={report}_report.csv"},
        )
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "report"
    for row in rows:
        ws.append(row)
    bio = io.BytesIO()
    wb.save(bio)
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={report}_report.xlsx"},
    )


@app.get("/admin/register-device", response_class=HTMLResponse)
def register_device_page(request: Request, db: Session = Depends(get_db)) -> HTMLResponse:
    employees = db.scalars(select(Employee).where(Employee.active.is_(True)).order_by(Employee.name)).all()
    return templates.TemplateResponse(
        "admin_register_device.html",
        {
            "request": request,
            "employees": employees,
            "register_link": "",
            "message": "",
        },
    )


@app.get("/admin-time/register-device", response_class=HTMLResponse)
def register_device_alias(request: Request, db: Session = Depends(get_db)) -> HTMLResponse:
    return register_device_page(request, db)


@app.post("/admin/register-device", response_class=HTMLResponse)
def register_device_create(
    request: Request,
    employee_id: int = Form(...),
    db: Session = Depends(get_db),
) -> HTMLResponse:
    employees = db.scalars(select(Employee).where(Employee.active.is_(True)).order_by(Employee.name)).all()
    employee = db.get(Employee, employee_id)
    if not employee or not employee.active:
        return templates.TemplateResponse(
            "admin_register_device.html",
            {
                "request": request,
                "employees": employees,
                "register_link": "",
                "message": "Geçersiz çalışan seçimi.",
            },
        )

    token = token_urlsafe(32)
    reg = RegistrationToken(
        employee_id=employee.id,
        token=token,
        used=False,
        created_at=now_berlin(),
    )


@app.post("/admin-time/register-device", response_class=HTMLResponse)
def register_device_create_alias(
    request: Request,
    employee_id: int = Form(...),
    db: Session = Depends(get_db),
) -> HTMLResponse:
    return register_device_create(request, employee_id, db)
    db.add(reg)
    db.commit()
    link = str(request.url_for("register_device_use")) + f"?token={token}"
    return templates.TemplateResponse(
        "admin_register_device.html",
        {
            "request": request,
            "employees": employees,
            "register_link": link,
            "message": f"{employee.name} için kayıt linki oluşturuldu.",
        },
    )


@app.get("/register-device", response_class=HTMLResponse, name="register_device_use")
def register_device_use(request: Request, token: str, db: Session = Depends(get_db)) -> HTMLResponse:
    reg = db.scalar(
        select(RegistrationToken)
        .options(joinedload(RegistrationToken.employee))
        .where(RegistrationToken.token == token)
    )
    if not reg:
        return templates.TemplateResponse(
            "register_status.html",
            {
                "request": request,
                "title": "Kayıt Başarısız",
                "message": "Kayıt bağlantısı geçersiz.",
                "ok": False,
            },
        )

    if reg.used:
        return templates.TemplateResponse(
            "register_status.html",
            {
                "request": request,
                "title": "Kayıt Kullanılmış",
                "message": "Bu kayıt bağlantısı daha önce kullanılmış.",
                "ok": False,
            },
        )

    device_token = token_urlsafe(32)
    device = Device(
        employee_id=reg.employee_id,
        device_token=device_token,
        created_at=now_berlin(),
        active=True,
    )
    reg.used = True
    db.add(device)
    db.commit()

    response = templates.TemplateResponse(
        "register_status.html",
        {
            "request": request,
            "title": "Cihaz Kaydı Tamamlandı",
            "message": "Cihaz başarıyla kaydedildi. Artık araç QR kodundan mesai sayfasına erişebilirsiniz.",
            "ok": True,
        },
    )
    response.set_cookie(
        key=DEVICE_COOKIE,
        value=device_token,
        httponly=True,
        secure=COOKIE_SECURE,
        samesite="lax",
        max_age=60 * 60 * 24 * 365,
    )
    return response


@app.get("/admin/vehicles", response_class=HTMLResponse)
def admin_vehicles(request: Request, db: Session = Depends(get_db)) -> HTMLResponse:
    vehicles = db.scalars(select(Vehicle).order_by(Vehicle.name)).all()
    base_url = str(request.base_url).rstrip("/")
    vehicle_qrs: list[dict] = []
    for v in vehicles:
        time_url = f"{base_url}/time?vehicle={v.qr_code_slug}"
        img = qrcode.make(time_url)
        buff = io.BytesIO()
        img.save(buff, format="PNG")
        b64 = base64.b64encode(buff.getvalue()).decode("ascii")
        vehicle_qrs.append(
            {
                "vehicle": v,
                "time_url": time_url,
                "qr_base64": b64,
            }
        )
    return templates.TemplateResponse(
        "admin_vehicles.html",
        {
            "request": request,
            "vehicle_qrs": vehicle_qrs,
        },
    )
