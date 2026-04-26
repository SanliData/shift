import hashlib
import hmac
import io
import logging
import re
from datetime import datetime, timedelta
from secrets import token_urlsafe
from urllib.parse import quote
from zoneinfo import ZoneInfo

import qrcode
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from fastapi.templating import Jinja2Templates
from sqlalchemy import desc, func, or_, select, text, update
from sqlalchemy.orm import Session, selectinload

from ..config import BASE_URL, COOKIE_SECURE, REGISTRATION_SIGNING_SECRET, TIMEZONE
from ..database import get_db
from ..models import (
    Device,
    Employee,
    ImportedFile,
    ProvisionalWorker,
    RegistrationToken,
    TimeEntry,
    TimeEntryCorrection,
    Vehicle,
)
from ..models import PW_STATUS_ACTIVE, PW_STATUS_DEACTIVATED, PW_STATUS_PENDING
from ..sqlite_migrations import ensure_provisional_schema, ensure_reporting_schema

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")
BERLIN_TZ = ZoneInfo(TIMEZONE)
DEVICE_COOKIE = "device_token"
BASE_TIME_URL = f"{BASE_URL}/time"
REG_TOKEN_MIGRATION_DONE = False
logger = logging.getLogger(__name__)
PREVIEW_UA_MARKERS = ("whatsapp", "facebookexternalhit", "meta", "telegrambot")


def now_berlin() -> datetime:
    return datetime.now(BERLIN_TZ)


def parse_date(date_str: str | None):
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=BERLIN_TZ)
    except ValueError:
        return None


def parse_datetime_local(raw: str | None) -> datetime | None:
    s = (raw or "").strip()
    if not s:
        return None
    try:
        s16 = s[:16].replace(" ", "T")
        dt = datetime.strptime(s16, "%Y-%m-%dT%H:%M")
        return dt.replace(tzinfo=BERLIN_TZ)
    except ValueError:
        return None


def fmt_datetime_local(dt: datetime | None) -> str:
    if not dt:
        return ""
    return as_berlin(dt).strftime("%Y-%m-%dT%H:%M")


def parse_optional_float_hours(raw: str | None) -> float | None:
    s = (raw or "").strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def recalc_minutes_and_costs(
    employee: Employee,
    *,
    start: datetime | None,
    end: datetime | None,
    normal_hours_override: float | None,
    overtime_hours_override: float | None,
) -> tuple[int, int, int, float, float, float]:
    """total_minutes, regular_minutes, overtime_minutes, regular_cost, overtime_cost, total_cost."""
    if normal_hours_override is not None and overtime_hours_override is not None:
        reg_m = max(0, int(round(normal_hours_override * 60)))
        ot_m = max(0, int(round(overtime_hours_override * 60)))
    elif start and end:
        delta = max(0, int((as_berlin(end) - as_berlin(start)).total_seconds() // 60))
        reg_m = min(delta, 480)
        ot_m = max(0, delta - 480)
    else:
        reg_m = 0
        ot_m = 0
    total_m = reg_m + ot_m
    hr = float(employee.hourly_rate or 0)
    otr = employee_overtime_rate(employee)
    reg_c = round((reg_m / 60.0) * hr, 2)
    ot_c = round((ot_m / 60.0) * otr, 2)
    tot_c = round(reg_c + ot_c, 2)
    return total_m, reg_m, ot_m, reg_c, ot_c, tot_c


def build_filters(date_from: str | None, date_to: str | None, employee_id: int | None, vehicle_id: int | None):
    filters = []
    start = parse_date(date_from)
    end = parse_date(date_to)
    if start:
        filters.append(TimeEntry.start_time >= start)
    if end:
        filters.append(TimeEntry.start_time <= end.replace(hour=23, minute=59, second=59))
    if employee_id:
        filters.append(TimeEntry.employee_id == employee_id)
    if vehicle_id:
        filters.append(TimeEntry.vehicle_id == vehicle_id)
    return filters


def to_slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.strip().lower()).strip("-")
    return slug


def as_berlin(dt: datetime) -> datetime:
    if dt.tzinfo is None:
        return dt.replace(tzinfo=BERLIN_TZ)
    return dt.astimezone(BERLIN_TZ)


def eur(value: float | int | None) -> str:
    return f"€{float(value or 0):,.2f}"


def hours_str(minutes: int | float | None) -> str:
    return f"{(float(minutes or 0) / 60):.2f}"


def correction_request_audit(request: Request) -> tuple[str, str, str | None, str | None]:
    """Placeholder identity until admin auth exists; store request fingerprint."""
    client_ip = request.client.host if request.client else None
    ua = request.headers.get("user-agent")
    if ua and len(ua) > 4000:
        ua = ua[:4000]
    return "admin", "admin", client_ip, ua


def time_entry_original_snapshot(
    entry: TimeEntry, employees: list[Employee], vehicles: list[Vehicle]
) -> dict[str, str]:
    emp_map = {e.id: e for e in employees}
    v_map = {v.id: v for v in vehicles}
    emp = emp_map.get(entry.employee_id) if entry.employee_id is not None else None
    v = v_map.get(entry.vehicle_id)
    return {
        "employee_label": emp.name if emp else entry.employee_name,
        "vehicle_label": v.name if v else f"Araç #{entry.vehicle_id}",
        "clock_in": as_berlin(entry.start_time).strftime("%d.%m.%Y %H:%M") if entry.start_time else "—",
        "clock_out": as_berlin(entry.end_time).strftime("%d.%m.%Y %H:%M") if entry.end_time else "—",
        "regular_hours": hours_str(entry.regular_minutes),
        "overtime_hours": hours_str(entry.overtime_minutes),
        "total_cost_eur": eur(entry.total_cost),
    }


def _time_entry_edit_page_context(
    request: Request,
    entry: TimeEntry,
    employees: list[Employee],
    vehicles: list[Vehicle],
    corrections: list[TimeEntryCorrection],
    *,
    clock_in_value: str,
    clock_out_value: str,
    normal_hours_value: str,
    overtime_hours_value: str,
    reason_value: str,
    form_error: str | None,
) -> dict[str, object]:
    return {
        "request": request,
        "berlin_now": now_berlin().strftime("%d.%m.%Y %H:%M:%S"),
        "entry": entry,
        "employees": employees,
        "vehicles": vehicles,
        "corrections": corrections,
        "original": time_entry_original_snapshot(entry, employees, vehicles),
        "clock_in_value": clock_in_value,
        "clock_out_value": clock_out_value,
        "normal_hours_value": normal_hours_value,
        "overtime_hours_value": overtime_hours_value,
        "reason_value": reason_value,
        "form_error": form_error,
    }


def _correction_row_display(c: TimeEntryCorrection) -> dict[str, object]:
    ua = (c.corrected_by_user_agent or "").replace("\n", " ").strip()
    if len(ua) > 140:
        ua = ua[:137] + "…"
    return {
        "created": as_berlin(c.created_at).strftime("%d.%m.%Y %H:%M"),
        "entry_id": c.time_entry_id,
        "by": c.corrected_by or "admin",
        "role": c.corrected_by_role or "admin",
        "ip": c.corrected_by_ip or "—",
        "ua": ua or "—",
        "reason": (c.reason or "").strip() or "—",
        "old_in": c.old_clock_in.strftime("%d.%m.%Y %H:%M") if c.old_clock_in else "—",
        "new_in": c.new_clock_in.strftime("%d.%m.%Y %H:%M") if c.new_clock_in else "—",
        "old_out": c.old_clock_out.strftime("%d.%m.%Y %H:%M") if c.old_clock_out else "—",
        "new_out": c.new_clock_out.strftime("%d.%m.%Y %H:%M") if c.new_clock_out else "—",
        "emp_ids": f"{c.old_employee_id} → {c.new_employee_id}",
        "veh_ids": f"{c.old_vehicle_id} → {c.new_vehicle_id}",
    }


def employee_overtime_rate(employee: Employee | None) -> float:
    if not employee:
        return 0.0
    if employee.overtime_hourly_rate is not None:
        return float(employee.overtime_hourly_rate or 0)
    return round(float(employee.hourly_rate or 0) * float(employee.overtime_multiplier or 1.5), 2)


def parse_optional_float(value: str | None) -> float | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    if cleaned == "":
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def normalize_phone_digits(phone: str | None) -> str:
    if not phone:
        return ""
    return re.sub(r"\D+", "", phone)


def build_register_link(token: str) -> str:
    return f"{BASE_URL}/register-device?token={token}"


def build_self_register_link() -> str:
    return f"{BASE_URL}/register-self"


def provisional_sign(provisional_id: int) -> str:
    return hmac.new(
        REGISTRATION_SIGNING_SECRET.encode("utf-8"),
        str(provisional_id).encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()[:48]


def build_whatsapp_link(phone_digits: str, register_link: str) -> str:
    text = f"Mesai sistemi cihaz kayit linkiniz: {register_link}"
    return f"https://wa.me/{phone_digits}?text={quote(text)}"


def build_sms_uri(phone_digits: str, register_link: str) -> str:
    body = f"Mesai sistemi cihaz kayit linkiniz: {register_link}"
    return f"sms:+{phone_digits}?body={quote(body)}"


def ensure_registration_token_columns(db: Session):
    global REG_TOKEN_MIGRATION_DONE
    if REG_TOKEN_MIGRATION_DONE:
        return
    cols = db.execute(text("PRAGMA table_info(registration_tokens)")).fetchall()
    col_names = {c[1] for c in cols}
    if cols and "active" not in col_names:
        db.execute(text("ALTER TABLE registration_tokens ADD COLUMN active BOOLEAN DEFAULT 1"))
        db.commit()
    if cols and "last_sent_at" not in col_names:
        db.execute(text("ALTER TABLE registration_tokens ADD COLUMN last_sent_at DATETIME"))
        db.commit()
    if cols:
        db.execute(text("UPDATE registration_tokens SET used = 0 WHERE used IS NULL"))
        db.execute(text("UPDATE registration_tokens SET active = 1 WHERE active IS NULL"))
        db.execute(text("UPDATE registration_tokens SET active = 0 WHERE used = 1"))
        employee_ids = db.execute(text("SELECT DISTINCT employee_id FROM registration_tokens")).fetchall()
        for row in employee_ids:
            employee_id = int(row[0])
            valid_rows = db.scalars(
                select(RegistrationToken)
                .where(
                    RegistrationToken.employee_id == employee_id,
                    RegistrationToken.active.is_(True),
                    RegistrationToken.used.is_(False),
                )
                .order_by(desc(RegistrationToken.created_at), desc(RegistrationToken.id))
            ).all()
            for old in valid_rows[1:]:
                old.active = False
        db.commit()
    REG_TOKEN_MIGRATION_DONE = True


def get_active_registration_token(db: Session, employee_id: int) -> RegistrationToken | None:
    rows = db.scalars(
        select(RegistrationToken)
        .where(
            RegistrationToken.employee_id == employee_id,
            RegistrationToken.active.is_(True),
            RegistrationToken.used.is_(False),
        )
        .order_by(desc(RegistrationToken.created_at), desc(RegistrationToken.id))
    ).all()
    if not rows:
        return None
    for old in rows[1:]:
        old.active = False
    return rows[0]


def get_latest_registration_token(db: Session, employee_id: int) -> RegistrationToken | None:
    return db.scalar(
        select(RegistrationToken)
        .where(RegistrationToken.employee_id == employee_id)
        .order_by(desc(RegistrationToken.created_at))
    )


def create_registration_token(db: Session, employee_id: int, *, deactivate_existing: bool) -> RegistrationToken:
    if deactivate_existing:
        tokens = db.scalars(
            select(RegistrationToken).where(
                RegistrationToken.employee_id == employee_id,
                RegistrationToken.active.is_(True),
            )
        ).all()
        for row in tokens:
            row.active = False
    token_value = token_urlsafe(24)
    new_row = RegistrationToken(
        employee_id=employee_id,
        token=token_value,
        active=True,
        used=False,
        created_at=now_berlin(),
    )
    db.add(new_row)
    db.flush()
    # Defensive write for legacy SQLite rows/default anomalies.
    new_row.used = False
    new_row.active = True
    logger.debug(
        "registration token created employee_id=%s token=%s used=%s active=%s",
        employee_id,
        token_value,
        new_row.used,
        new_row.active,
    )
    return new_row


def get_or_create_valid_registration_token(
    db: Session,
    employee_id: int,
    *,
    regenerate: bool = False,
) -> RegistrationToken | None:
    employee = db.scalar(select(Employee).where(Employee.id == employee_id))
    if not employee or not employee.active:
        return None
    token_row = get_active_registration_token(db, employee_id)
    if regenerate or not token_row:
        token_row = create_registration_token(db, employee_id, deactivate_existing=True)
    return token_row


def get_registerable_token(db: Session, raw_token: str) -> RegistrationToken | None:
    clean_token = (raw_token or "").strip()
    return db.scalar(
        select(RegistrationToken)
        .join(Employee, Employee.id == RegistrationToken.employee_id)
        .where(
            RegistrationToken.token == clean_token,
            RegistrationToken.used.is_(False),
            RegistrationToken.active.is_(True),
            Employee.active.is_(True),
        )
    )


def is_preview_request(request: Request) -> bool:
    if request.method.upper() == "HEAD":
        return True
    ua = (request.headers.get("user-agent") or "").lower()
    return any(marker in ua for marker in PREVIEW_UA_MARKERS)


def render_admin_time(
    request: Request,
    db: Session,
    *,
    date_from: str = "",
    date_to: str = "",
    employee_id: int | None = None,
    vehicle_id: int | None = None,
    message: str = "",
):
    ensure_registration_token_columns(db)
    ensure_provisional_schema(db)
    ensure_reporting_schema(db)
    employees = db.scalars(select(Employee).order_by(Employee.name)).all()
    vehicles = db.scalars(select(Vehicle).order_by(Vehicle.qr_code_slug)).all()
    devices = db.scalars(
        select(Device)
        .options(selectinload(Device.employee), selectinload(Device.provisional_worker))
        .order_by(desc(Device.created_at))
    ).all()
    device_counts: dict[int, int] = {}
    for d in devices:
        if d.active and d.employee_id is not None:
            device_counts[d.employee_id] = device_counts.get(d.employee_id, 0) + 1

    filters = build_filters(date_from or None, date_to or None, employee_id, vehicle_id)
    active_stmt = select(TimeEntry).where(TimeEntry.status == "active").order_by(desc(TimeEntry.start_time))
    completed_stmt = select(TimeEntry).where(TimeEntry.status == "completed").order_by(desc(TimeEntry.end_time)).limit(200)
    for f in filters:
        active_stmt = active_stmt.where(f)
        completed_stmt = completed_stmt.where(f)
    active_entries = db.scalars(active_stmt).all()
    completed_entries = db.scalars(completed_stmt).all()
    employee_active_map: dict[int, int] = {}
    employee_total_minutes_map: dict[int, int] = {}
    for row in active_entries:
        if row.employee_id is not None:
            employee_active_map[row.employee_id] = employee_active_map.get(row.employee_id, 0) + 1
    all_completed = db.scalars(select(TimeEntry).where(TimeEntry.status == "completed")).all()
    for row in all_completed:
        if row.employee_id is not None:
            employee_total_minutes_map[row.employee_id] = employee_total_minutes_map.get(row.employee_id, 0) + int(
                row.total_minutes or 0
            )

    provisional_workers_visible = db.scalars(
        select(ProvisionalWorker)
        .where(ProvisionalWorker.status != PW_STATUS_DEACTIVATED)
        .order_by(desc(ProvisionalWorker.created_at))
    ).all()
    provisional_dashboard: list[dict] = []
    for pw in provisional_workers_visible:
        mins = db.scalar(
            select(func.coalesce(func.sum(TimeEntry.total_minutes), 0)).where(
                TimeEntry.provisional_worker_id == pw.id,
                TimeEntry.status == "completed",
            )
        )
        provisional_dashboard.append(
            {
                "id": pw.id,
                "full_name": pw.full_name,
                "phone": pw.phone,
                "date_of_birth": pw.date_of_birth or "—",
                "status": pw.status,
                "minutes": int(mins or 0),
                "can_approve": pw.status == PW_STATUS_ACTIVE,
            }
        )

    now = now_berlin()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=today_start.weekday())
    month_start = today_start.replace(day=1)
    summary_entries = db.scalars(select(TimeEntry).where(TimeEntry.status == "completed")).all()
    today_minutes = 0
    week_minutes = 0
    month_minutes = 0
    today_cost = 0.0
    week_cost = 0.0
    month_cost = 0.0
    month_overtime_cost = 0.0
    for row in summary_entries:
        if not row.start_time:
            continue
        local_start = as_berlin(row.start_time)
        minutes = int(row.total_minutes or 0)
        row_total_cost = float(row.total_cost or 0)
        row_overtime_cost = float(row.overtime_cost or 0)
        if local_start >= today_start:
            today_minutes += minutes
            today_cost += row_total_cost
        if local_start >= week_start:
            week_minutes += minutes
            week_cost += row_total_cost
        if local_start >= month_start:
            month_minutes += minutes
            month_cost += row_total_cost
            month_overtime_cost += row_overtime_cost
    missing_rate_employees = [e.name for e in employees if float(e.hourly_rate or 0) <= 0]
    employee_actions: dict[int, dict[str, str]] = {}
    for emp in employees:
        token_row = get_active_registration_token(db, emp.id)
        latest_token = get_latest_registration_token(db, emp.id)
        if not token_row:
            if latest_token and latest_token.used:
                token_status = "Kullanılmış"
            else:
                token_status = "Token yok"
            last_sent = as_berlin(latest_token.last_sent_at).strftime("%d.%m.%Y %H:%M") if (latest_token and latest_token.last_sent_at) else "-"
            employee_actions[emp.id] = {
                "register_link": "",
                "wa_url": "",
                "sms_uri": "",
                "phone_digits": "",
                "token_status": token_status,
                "token_active": False,
                "last_sent_at": last_sent,
                "resend_wa_url": f"/admin-time/employees/{emp.id}/device-link?channel=whatsapp",
                "resend_sms_url": f"/admin-time/employees/{emp.id}/device-link?channel=sms",
                "link_url": f"/admin-time/employees/{emp.id}/device-link",
                "regenerate_url": f"/admin-time/employees/{emp.id}/regenerate-link",
            }
            continue
        register_link_value = build_register_link(token_row.token)
        digits = normalize_phone_digits(emp.phone_number)
        wa_url = build_whatsapp_link(digits, register_link_value) if digits else ""
        sms_uri = build_sms_uri(digits, register_link_value) if digits else ""
        last_sent = as_berlin(token_row.last_sent_at).strftime("%d.%m.%Y %H:%M") if token_row.last_sent_at else "-"
        employee_actions[emp.id] = {
            "register_link": register_link_value,
            "wa_url": wa_url,
            "sms_uri": sms_uri,
            "phone_digits": digits,
            "token_status": "Geçerli token",
            "token_active": True,
            "last_sent_at": last_sent,
            "resend_wa_url": f"/admin-time/employees/{emp.id}/device-link?channel=whatsapp",
            "resend_sms_url": f"/admin-time/employees/{emp.id}/device-link?channel=sms",
            "link_url": f"/admin-time/employees/{emp.id}/device-link",
            "regenerate_url": f"/admin-time/employees/{emp.id}/regenerate-link",
        }
    totals = db.execute(
        select(
            func.coalesce(func.sum(TimeEntry.total_minutes), 0),
            func.coalesce(func.sum(TimeEntry.overtime_minutes), 0),
        ).where(TimeEntry.status == "completed")
    ).one()
    return templates.TemplateResponse(
        request=request,
        name="admin_time.html",
        context={
            "request": request,
            "employees": employees,
            "device_counts": device_counts,
            "vehicles": vehicles,
            "devices": devices,
            "active_entries": active_entries,
            "completed_entries": completed_entries,
            "employee_active_map": employee_active_map,
            "employee_total_minutes_map": employee_total_minutes_map,
            "total_minutes": int(totals[0] or 0),
            "overtime_minutes": int(totals[1] or 0),
            "today_hours": round(today_minutes / 60, 2),
            "week_hours": round(week_minutes / 60, 2),
            "month_hours": round(month_minutes / 60, 2),
            "active_shift_count": len(active_entries),
            "today_cost_eur": eur(today_cost),
            "week_cost_eur": eur(week_cost),
            "month_cost_eur": eur(month_cost),
            "month_overtime_cost_eur": eur(month_overtime_cost),
            "missing_rate_employees": missing_rate_employees,
            "employee_actions": employee_actions,
            "berlin_now": now.strftime("%d.%m.%Y %H:%M:%S"),
            "filters": {
                "date_from": date_from or "",
                "date_to": date_to or "",
                "employee_id": employee_id or "",
                "vehicle_id": vehicle_id or "",
            },
            "message": message,
            "self_register_link": build_self_register_link(),
            "provisional_dashboard": provisional_dashboard,
        },
    )


@router.get("/admin-time", response_class=HTMLResponse)
def admin_page(
    request: Request,
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    employee_id: int | None = Query(default=None),
    vehicle_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
):
    return render_admin_time(
        request,
        db,
        date_from=date_from or "",
        date_to=date_to or "",
        employee_id=employee_id,
        vehicle_id=vehicle_id,
    )


@router.post("/admin-time/employees", response_class=HTMLResponse)
def create_employee(
    request: Request,
    full_name: str = Form(...),
    phone_number: str = Form(...),
    hourly_rate: float = Form(0),
    overtime_hourly_rate: str = Form(""),
    overtime_multiplier: float = Form(1.5),
    active: str = Form("true"),
    db: Session = Depends(get_db),
):
    name = full_name.strip()
    phone = phone_number.strip()
    if not name:
        return render_admin_time(request, db, message="Ad-soyad boş olamaz.")
    exists = db.scalar(select(Employee).where(Employee.name == name, Employee.phone_number == phone))
    if exists:
        return render_admin_time(request, db, message="Bu çalışan zaten kayıtlı.")
    is_active = str(active).lower() in ("1", "true", "yes", "on")
    db.add(
        Employee(
            name=name,
            phone_number=phone,
            hourly_rate=max(0, float(hourly_rate or 0)),
            overtime_hourly_rate=(
                max(0, parse_optional_float(overtime_hourly_rate))
                if parse_optional_float(overtime_hourly_rate) is not None
                else None
            ),
            overtime_multiplier=max(1.0, float(overtime_multiplier or 1.5)),
            active=is_active,
        )
    )
    db.commit()
    return render_admin_time(request, db, message=f"{name} için çalışan kaydı oluşturuldu.")


@router.post("/admin-time/employees/{employee_id}/update", response_class=HTMLResponse)
def update_employee(
    request: Request,
    employee_id: int,
    full_name: str = Form(...),
    phone_number: str = Form(""),
    hourly_rate: str = Form(""),
    overtime_hourly_rate: str = Form(""),
    overtime_multiplier: float = Form(1.5),
    active: str = Form("true"),
    db: Session = Depends(get_db),
):
    employee = db.scalar(select(Employee).where(Employee.id == employee_id))
    if not employee:
        return render_admin_time(request, db, message="Çalışan bulunamadı.")
    name = full_name.strip()
    if not name:
        return render_admin_time(request, db, message="Ad soyad boş olamaz.")
    employee.name = name
    employee.phone_number = phone_number.strip() or None
    hourly_parsed = parse_optional_float(hourly_rate.strip() if hourly_rate else "")
    employee.hourly_rate = max(0.0, float(hourly_parsed if hourly_parsed is not None else 0.0))
    parsed_ot_rate = parse_optional_float(overtime_hourly_rate.strip() if overtime_hourly_rate else "")
    employee.overtime_hourly_rate = max(0, parsed_ot_rate) if parsed_ot_rate is not None else None
    employee.overtime_multiplier = max(1.0, float(overtime_multiplier or 1.5))
    employee.active = str(active).lower() in ("1", "true", "yes", "on")
    db.commit()
    return render_admin_time(request, db, message="Çalışan bilgileri güncellendi.")


@router.post("/admin-time/register-link")
def create_register_link_deprecated():
    return JSONResponse(
        {"detail": "Deprecated. Use GET /admin-time/employees/{employee_id}/device-link"},
        status_code=410,
    )


@router.get("/admin-time/employees/{employee_id}/device-link")
def employee_device_link(
    employee_id: int,
    channel: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    ensure_registration_token_columns(db)
    channel_name = (channel or "").lower()
    employee = db.scalar(select(Employee).where(Employee.id == employee_id))
    if not employee:
        if channel_name in ("whatsapp", "sms"):
            return RedirectResponse("/admin-time?message=Çalışan+bulunamadı.", status_code=303)
        return JSONResponse({"detail": "Çalışan bulunamadı."}, status_code=404)
    token_row = get_or_create_valid_registration_token(db, employee_id)
    if not token_row:
        if channel_name in ("whatsapp", "sms"):
            return RedirectResponse("/admin-time?message=Çalışan+pasif.", status_code=303)
        return JSONResponse({"detail": "Çalışan pasif."}, status_code=400)
    register_link = build_register_link(token_row.token)
    phone_digits = normalize_phone_digits(employee.phone_number)
    if channel_name in ("whatsapp", "sms"):
        if not phone_digits:
            db.commit()
            return RedirectResponse("/admin-time?message=Telefon+numarası+gerekli.", status_code=303)
        token_row.last_sent_at = now_berlin()
        db.commit()
        if channel_name == "whatsapp":
            return RedirectResponse(build_whatsapp_link(phone_digits, register_link), status_code=302)
        return RedirectResponse(build_sms_uri(phone_digits, register_link), status_code=302)
    db.commit()
    return {
        "employee_id": employee_id,
        "token": token_row.token,
        "active": bool(token_row.active and (not token_row.used)),
        "used": bool(token_row.used),
        "register_link": register_link,
        "created_at": token_row.created_at.isoformat() if token_row.created_at else None,
        "last_sent_at": token_row.last_sent_at.isoformat() if token_row.last_sent_at else None,
    }


@router.post("/admin-time/employees/{employee_id}/regenerate-link")
def regenerate_employee_link(
    employee_id: int,
    request: Request,
    channel: str = Form(default=""),
    db: Session = Depends(get_db),
):
    ensure_registration_token_columns(db)
    employee = db.scalar(select(Employee).where(Employee.id == employee_id))
    if not employee:
        return render_admin_time(request, db, message="Çalışan bulunamadı.")
    row = get_or_create_valid_registration_token(db, employee_id, regenerate=True)
    if not row:
        return render_admin_time(request, db, message="Çalışan pasif.")
    register_link = build_register_link(row.token)
    phone_digits = normalize_phone_digits(employee.phone_number)
    channel_name = (channel or "").lower()
    if channel_name in ("whatsapp", "sms"):
        if not phone_digits:
            db.commit()
            return render_admin_time(
                request,
                db,
                message="Telefon numarası yok. Önce telefon ekleyin.",
            )
        row.last_sent_at = now_berlin()
        db.commit()
        if channel_name == "whatsapp":
            return RedirectResponse(build_whatsapp_link(phone_digits, register_link), status_code=302)
        return RedirectResponse(build_sms_uri(phone_digits, register_link), status_code=302)
    db.commit()
    return render_admin_time(request, db, message="Yeni kayıt linki üretildi.")


@router.get("/register-device", response_class=HTMLResponse)
def register_device(request: Request, token: str, db: Session = Depends(get_db)):
    ensure_registration_token_columns(db)
    clean_token = (token or "").strip()
    logger.debug("register-device request token=%s", clean_token)
    reg = get_registerable_token(db, clean_token)
    if is_preview_request(request):
        logger.debug("register-device preview request method=%s ua=%s", request.method, request.headers.get("user-agent", ""))
        return HTMLResponse("<html><body>Kayıt linki hazır. Lütfen linke tıklayın.</body></html>", status_code=200)
    if not reg:
        debug_row = db.scalar(select(RegistrationToken).where(RegistrationToken.token == clean_token))
        debug_employee_id = debug_row.employee_id if debug_row else None
        debug_used = debug_row.used if debug_row else None
        debug_active = debug_row.active if debug_row else None
        debug_emp_active = None
        if debug_employee_id:
            dbg_emp = db.scalar(select(Employee).where(Employee.id == debug_employee_id))
            debug_emp_active = dbg_emp.active if dbg_emp else None
        reason = "not_found"
        if not clean_token:
            reason = "empty_token"
        elif debug_row and debug_row.used:
            reason = "used_true"
        elif debug_row and not debug_row.active:
            reason = "token_not_active"
        elif debug_row and debug_emp_active is False:
            reason = "employee_inactive"
        logger.debug(
            "register-device rejected token=%s found=%s used=%s active=%s employee_id=%s employee_active=%s reason=%s",
            clean_token,
            bool(debug_row),
            debug_used,
            debug_active,
            debug_employee_id,
            debug_emp_active,
            reason,
        )
        return templates.TemplateResponse(
            request=request,
            name="register_status.html",
            context={"request": request, "ok": False, "message": "Geçersiz veya kullanılmış token."},
        )
    employee = db.scalar(select(Employee).where(Employee.id == reg.employee_id))
    return templates.TemplateResponse(
        request=request,
        name="register_status.html",
        context={
            "request": request,
            "ok": True,
            "awaiting_confirm": True,
            "message": "Kayıt linki hazır. Cihazı kaydet butonuna basın.",
            "token": clean_token,
            "employee_name": employee.name if employee else "",
            "employee_phone": employee.phone_number if employee else "",
        },
    )


@router.post("/register-device/confirm", response_class=HTMLResponse)
def register_device_confirm(request: Request, token: str = Form(...), db: Session = Depends(get_db)):
    ensure_registration_token_columns(db)
    clean_token = (token or "").strip()
    logger.debug("register-device confirm token=%s", clean_token)
    reg = get_registerable_token(db, clean_token)
    if not reg:
        return templates.TemplateResponse(
            request=request,
            name="register_status.html",
            context={"request": request, "ok": False, "message": "Geçersiz veya kullanılmış token."},
        )
    new_token = token_urlsafe(32)
    db.add(
        Device(
            employee_id=reg.employee_id,
            provisional_worker_id=None,
            device_token=new_token,
            created_at=now_berlin(),
            active=True,
        )
    )
    reg.used = True
    reg.active = False
    logger.debug(
        "register-device success token=%s employee_id=%s used=%s active=%s",
        clean_token,
        reg.employee_id,
        reg.used,
        reg.active,
    )
    db.commit()
    employee = db.scalar(select(Employee).where(Employee.id == reg.employee_id))
    default_vehicle = db.scalar(select(Vehicle).where(Vehicle.active.is_(True)).order_by(Vehicle.qr_code_slug))
    time_vehicle = default_vehicle.qr_code_slug if default_vehicle else "vehicle-01"
    time_entry_url = f"/time?vehicle={time_vehicle}"
    resp = templates.TemplateResponse(
        request=request,
        name="register_status.html",
        context={
            "request": request,
            "ok": True,
            "awaiting_confirm": False,
            "message": "Cihaz başarıyla kaydedildi.",
            "employee_name": employee.name if employee else "",
            "employee_phone": employee.phone_number if employee else "",
            "time_entry_url": time_entry_url,
        },
    )
    resp.set_cookie(
        DEVICE_COOKIE,
        new_token,
        httponly=True,
        secure=COOKIE_SECURE and request.url.scheme == "https",
        samesite="lax",
        max_age=31536000,
        path="/",
    )
    return resp


@router.get("/register-self", response_class=HTMLResponse)
def register_self_form(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="register_self.html",
        context={"request": request},
    )


@router.post("/register-self/start", response_class=HTMLResponse)
def register_self_start(
    request: Request,
    full_name: str = Form(...),
    phone: str = Form(...),
    date_of_birth: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_provisional_schema(db)
    name = (full_name or "").strip()
    ph = (phone or "").strip()
    if not name or not ph:
        return templates.TemplateResponse(
            request=request,
            name="register_self.html",
            context={"request": request, "error": "Ad soyad ve telefon zorunludur."},
        )
    dob = (date_of_birth or "").strip() or None
    pw = ProvisionalWorker(
        full_name=name,
        phone=ph,
        date_of_birth=dob,
        device_token=None,
        created_at=now_berlin(),
        status=PW_STATUS_PENDING,
    )
    db.add(pw)
    db.commit()
    db.refresh(pw)
    key = provisional_sign(pw.id)
    return RedirectResponse(url=f"/register-self/device?pid={pw.id}&key={key}", status_code=303)


@router.get("/register-self/device", response_class=HTMLResponse)
def register_self_device_page(request: Request, pid: int, key: str, db: Session = Depends(get_db)):
    ensure_provisional_schema(db)
    if is_preview_request(request):
        return HTMLResponse("<html><body>Kayıt linki hazır. Lütfen linke tıklayın.</body></html>", status_code=200)
    pw = db.scalar(select(ProvisionalWorker).where(ProvisionalWorker.id == pid))
    if not pw or pw.status != PW_STATUS_PENDING:
        return templates.TemplateResponse(
            request=request,
            name="register_self_device.html",
            context={"request": request, "ok": False, "message": "Kayıt bulunamadı veya zaten tamamlanmış."},
        )
    if not hmac.compare_digest(provisional_sign(pid), (key or "").strip()):
        return templates.TemplateResponse(
            request=request,
            name="register_self_device.html",
            context={"request": request, "ok": False, "message": "Geçersiz doğrulama bağlantısı."},
        )
    return templates.TemplateResponse(
        request=request,
        name="register_self_device.html",
        context={
            "request": request,
            "ok": True,
            "awaiting_confirm": True,
            "pid": pid,
            "key": key,
            "full_name": pw.full_name,
            "phone": pw.phone,
            "date_of_birth": pw.date_of_birth or "—",
        },
    )


@router.post("/register-self/confirm", response_class=HTMLResponse)
def register_self_confirm(
    request: Request,
    pid: int = Form(...),
    key: str = Form(...),
    db: Session = Depends(get_db),
):
    ensure_provisional_schema(db)
    pw = db.scalar(select(ProvisionalWorker).where(ProvisionalWorker.id == pid))
    if not pw or pw.status != PW_STATUS_PENDING:
        return templates.TemplateResponse(
            request=request,
            name="register_self_device.html",
            context={"request": request, "ok": False, "message": "Kayıt bulunamadı veya zaten tamamlanmış."},
        )
    if not hmac.compare_digest(provisional_sign(pid), (key or "").strip()):
        return templates.TemplateResponse(
            request=request,
            name="register_self_device.html",
            context={"request": request, "ok": False, "message": "Geçersiz doğrulama."},
        )
    existing = db.scalar(
        select(Device).where(Device.provisional_worker_id == pid, Device.active.is_(True))
    )
    if existing:
        return templates.TemplateResponse(
            request=request,
            name="register_self_device.html",
            context={"request": request, "ok": False, "message": "Bu kayıt için cihaz zaten tanımlı."},
        )
    new_token = token_urlsafe(32)
    db.add(
        Device(
            employee_id=None,
            provisional_worker_id=pw.id,
            device_token=new_token,
            created_at=now_berlin(),
            active=True,
        )
    )
    pw.device_token = new_token
    pw.status = PW_STATUS_ACTIVE
    db.commit()
    default_vehicle = db.scalar(select(Vehicle).where(Vehicle.active.is_(True)).order_by(Vehicle.qr_code_slug))
    time_vehicle = default_vehicle.qr_code_slug if default_vehicle else "vehicle-01"
    time_entry_url = f"/time?vehicle={time_vehicle}"
    resp = templates.TemplateResponse(
        request=request,
        name="register_self_device.html",
        context={
            "request": request,
            "ok": True,
            "awaiting_confirm": False,
            "message": "Cihaz kaydedildi. Mesaiye başlayabilirsiniz.",
            "full_name": pw.full_name,
            "phone": pw.phone,
            "date_of_birth": pw.date_of_birth or "—",
            "time_entry_url": time_entry_url,
        },
    )
    resp.set_cookie(
        DEVICE_COOKIE,
        new_token,
        httponly=True,
        secure=COOKIE_SECURE and request.url.scheme == "https",
        samesite="lax",
        max_age=31536000,
        path="/",
    )
    return resp


@router.post("/admin-time/provisional-workers/{pw_id}/approve", response_class=HTMLResponse)
def approve_provisional_worker(
    pw_id: int,
    request: Request,
    hourly_rate: float = Form(20.0),
    overtime_hourly_rate: str = Form(""),
    overtime_multiplier: float = Form(1.5),
    active: str = Form("true"),
    db: Session = Depends(get_db),
):
    ensure_provisional_schema(db)
    pw = db.scalar(select(ProvisionalWorker).where(ProvisionalWorker.id == pw_id))
    if not pw or pw.status != PW_STATUS_ACTIVE:
        return render_admin_time(request, db, message="Ön kayıt bulunamadı veya onaylanamaz durumda.")
    digits = normalize_phone_digits(pw.phone)
    if digits:
        conflict = next(
            (
                e
                for e in db.scalars(select(Employee)).all()
                if normalize_phone_digits(e.phone_number) == digits
            ),
            None,
        )
        if conflict:
            return render_admin_time(
                request,
                db,
                message="Bu telefon numarasıyla kayıtlı çalışan zaten var. Önce mevcut kaydı düzenleyin.",
            )
    ot_parsed = parse_optional_float(overtime_hourly_rate)
    is_active = str(active).lower() in ("1", "true", "yes", "on")
    emp = Employee(
        name=pw.full_name.strip(),
        phone_number=pw.phone.strip() or None,
        hourly_rate=max(0.0, float(hourly_rate or 0)),
        overtime_multiplier=max(1.0, float(overtime_multiplier or 1.5)),
        overtime_hourly_rate=(max(0.0, ot_parsed) if ot_parsed is not None else None),
        active=is_active,
    )
    db.add(emp)
    db.flush()

    db.execute(
        update(TimeEntry)
        .where(TimeEntry.provisional_worker_id == pw_id)
        .values(employee_id=emp.id, employee_name=emp.name, provisional_worker_id=None)
    )
    for dev in db.scalars(select(Device).where(Device.provisional_worker_id == pw_id)).all():
        dev.employee_id = emp.id
        dev.provisional_worker_id = None
    pw.status = PW_STATUS_DEACTIVATED
    db.commit()
    return render_admin_time(request, db, message=f"{emp.name} çalışan olarak oluşturuldu; mesai kayıtları taşındı.")


@router.get("/admin-time/import", response_class=HTMLResponse)
def import_page(request: Request, db: Session = Depends(get_db)):
    files = db.scalars(select(ImportedFile).order_by(desc(ImportedFile.created_at)).limit(20)).all()
    return templates.TemplateResponse(
        request=request,
        name="import.html",
        context={"request": request, "files": files, "message": "", "error": ""},
    )


@router.post("/admin-time/import", response_class=HTMLResponse)
async def import_upload(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    data = await file.read()
    imported_rows = 0
    error = ""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), data_only=True)
        if "VERI_GIRISI" in wb.sheetnames:
            ws = wb["VERI_GIRISI"]
            imported_rows = max(0, ws.max_row - 1)
        else:
            imported_rows = 0
        db.add(ImportedFile(filename=file.filename or "uploaded.xlsx", imported_rows=imported_rows, created_at=now_berlin()))
        db.commit()
    except Exception as ex:
        error = str(ex)
    files = db.scalars(select(ImportedFile).order_by(desc(ImportedFile.created_at)).limit(20)).all()
    return templates.TemplateResponse(
        request=request,
        name="import.html",
        context={"request": request, "files": files, "message": f"Import tamamlandı. Satır: {imported_rows}", "error": error},
    )


@router.get("/admin-time/reports", response_class=HTMLResponse)
def admin_time_reports(
    request: Request,
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    employee_id: int | None = Query(default=None),
    vehicle_id: int | None = Query(default=None),
    month: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    ensure_reporting_schema(db)
    employees = db.scalars(select(Employee).order_by(Employee.name)).all()
    vehicles = db.scalars(select(Vehicle).order_by(Vehicle.name)).all()
    employee_map = {e.id: e for e in employees}
    vehicle_map = {v.id: v for v in vehicles}
    filters = build_filters(date_from, date_to, employee_id, vehicle_id)
    entries_stmt = select(TimeEntry).order_by(desc(TimeEntry.start_time))
    for f in filters:
        entries_stmt = entries_stmt.where(f)
    entries = db.scalars(entries_stmt).all()
    if month:
        try:
            year_s, month_s = month.split("-")
            year_i, month_i = int(year_s), int(month_s)
            entries = [e for e in entries if e.start_time and as_berlin(e.start_time).year == year_i and as_berlin(e.start_time).month == month_i]
        except Exception:
            pass

    entries = [e for e in entries if e.status == "completed"]

    employee_report: dict[int, dict[str, object]] = {}
    vehicle_report: dict[int, dict[str, object]] = {}
    daily_rows: list[dict] = []
    for e in entries:
        emp = employee_map.get(e.employee_id) if e.employee_id is not None else None
        hourly_rate = float(emp.hourly_rate or 0) if emp else 0.0
        overtime_rate = employee_overtime_rate(emp)
        regular_minutes = int(e.regular_minutes or 0)
        overtime_minutes = int(e.overtime_minutes or 0)
        total_minutes = int(e.total_minutes or 0)
        regular_cost = float(e.regular_cost or 0)
        overtime_cost = float(e.overtime_cost or 0)
        total_cost = float(e.total_cost or 0)

        if e.employee_id is not None:
            if e.employee_id not in employee_report:
                employee_report[e.employee_id] = {
                    "employee_name": e.employee_name,
                    "phone": (emp.phone_number if emp else None) or "-",
                    "days": set(),
                    "regular_minutes": 0,
                    "overtime_minutes": 0,
                    "total_minutes": 0,
                    "hourly_rate": hourly_rate,
                    "overtime_hourly_rate": overtime_rate,
                    "regular_cost": 0.0,
                    "overtime_cost": 0.0,
                    "total_cost": 0.0,
                    "missing_rate": hourly_rate <= 0,
                }
            r = employee_report[e.employee_id]
            r["days"].add(as_berlin(e.start_time).strftime("%Y-%m-%d") if e.start_time else "-")
            r["regular_minutes"] += regular_minutes
            r["overtime_minutes"] += overtime_minutes
            r["total_minutes"] += total_minutes
            r["regular_cost"] += regular_cost
            r["overtime_cost"] += overtime_cost
            r["total_cost"] += total_cost

        if e.vehicle_id not in vehicle_report:
            vehicle_report[e.vehicle_id] = {
                "vehicle_name": vehicle_map[e.vehicle_id].name if e.vehicle_id in vehicle_map else f"ID {e.vehicle_id}",
                "vehicle_type": (vehicle_map[e.vehicle_id].type if e.vehicle_id in vehicle_map else "-") or "-",
                "total_hours": 0.0,
                "employee_ids": set(),
                "total_cost": 0.0,
            }
        vr = vehicle_report[e.vehicle_id]
        vr["total_hours"] += float(total_minutes) / 60
        if e.employee_id is not None:
            vr["employee_ids"].add(e.employee_id)
        vr["total_cost"] += total_cost

        daily_rows.append(
            {
                "entry_id": e.id,
                "employee_id": e.employee_id,
                "vehicle_id": e.vehicle_id,
                "date": as_berlin(e.start_time).strftime("%Y-%m-%d") if e.start_time else "-",
                "employee_name": e.employee_name,
                "phone": (emp.phone_number if emp else None) or "-",
                "vehicle_name": vehicle_map[e.vehicle_id].name if e.vehicle_id in vehicle_map else f"ID {e.vehicle_id}",
                "start_time": as_berlin(e.start_time).strftime("%d.%m.%Y %H:%M") if e.start_time else "-",
                "end_time": as_berlin(e.end_time).strftime("%d.%m.%Y %H:%M") if e.end_time else "-",
                "regular_hours": f"{(regular_minutes / 60):.2f}",
                "overtime_hours": f"{(overtime_minutes / 60):.2f}",
                "total_hours": f"{(total_minutes / 60):.2f}",
                "regular_cost_eur": eur(regular_cost),
                "overtime_cost_eur": eur(overtime_cost),
                "total_cost_eur": eur(total_cost),
            }
        )
    vehicle_rows = []
    for vid in sorted(vehicle_report.keys(), key=lambda i: str(vehicle_report[i]["vehicle_name"])):
        row = vehicle_report[vid]
        vehicle_rows.append(
            {
                "vehicle_id": vid,
                "vehicle_name": row["vehicle_name"],
                "vehicle_type": row["vehicle_type"],
                "total_hours": round(row["total_hours"], 2),
                "employee_count": len(row["employee_ids"]),
                "total_cost_eur": eur(row["total_cost"]),
            }
        )
    employee_rows = []
    for eid in sorted(employee_report.keys(), key=lambda i: str(employee_report[i]["employee_name"])):
        row = employee_report[eid]
        employee_rows.append(
            {
                "employee_id": eid,
                "employee_name": row["employee_name"],
                "phone": row["phone"],
                "worked_days": len(row["days"]),
                "regular_hours": f"{(row['regular_minutes'] / 60):.2f}",
                "overtime_hours": f"{(row['overtime_minutes'] / 60):.2f}",
                "total_hours": f"{(row['total_minutes'] / 60):.2f}",
                "hourly_rate_eur": eur(row["hourly_rate"]),
                "overtime_hourly_rate_eur": eur(row["overtime_hourly_rate"]),
                "regular_cost_eur": eur(row["regular_cost"]),
                "overtime_cost_eur": eur(row["overtime_cost"]),
                "total_cost_eur": eur(row["total_cost"]),
                "missing_rate": row["missing_rate"],
            }
        )
    return templates.TemplateResponse(
        request=request,
        name="admin_time_reports.html",
        context={
            "request": request,
            "berlin_now": now_berlin().strftime("%d.%m.%Y %H:%M:%S"),
            "employees": employees,
            "vehicles": vehicles,
            "employee_rows": employee_rows,
            "vehicle_rows": vehicle_rows,
            "daily_rows": daily_rows,
            "filters": {
                "date_from": date_from or "",
                "date_to": date_to or "",
                "employee_id": employee_id or "",
                "vehicle_id": vehicle_id or "",
                "month": month or "",
            },
        },
    )


@router.get("/admin-time/dashboard")
def admin_time_dashboard_redirect():
    return RedirectResponse("/admin-time/reports", status_code=302)


@router.get("/admin-time/employees/{employee_id}/profile", response_class=HTMLResponse)
def admin_employee_profile(request: Request, employee_id: int, db: Session = Depends(get_db)):
    ensure_reporting_schema(db)
    employee = db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    entries = db.scalars(
        select(TimeEntry)
        .where(TimeEntry.employee_id == employee_id, TimeEntry.status == "completed")
        .order_by(desc(TimeEntry.start_time))
    ).all()
    vehicle_map = {v.id: v for v in db.scalars(select(Vehicle)).all()}
    day_set: set[str] = set()
    total_reg_m = 0
    total_ot_m = 0
    total_cost = 0.0
    history: list[dict[str, object]] = []
    for e in entries:
        if e.start_time:
            day_set.add(as_berlin(e.start_time).strftime("%Y-%m-%d"))
        reg_m = int(e.regular_minutes or 0)
        ot_m = int(e.overtime_minutes or 0)
        tot_m = int(e.total_minutes or 0)
        total_reg_m += reg_m
        total_ot_m += ot_m
        total_cost += float(e.total_cost or 0)
        vname = vehicle_map[e.vehicle_id].name if e.vehicle_id in vehicle_map else f"ID {e.vehicle_id}"
        history.append(
            {
                "entry_id": e.id,
                "date": as_berlin(e.start_time).strftime("%Y-%m-%d") if e.start_time else "-",
                "vehicle_name": vname,
                "start_time": as_berlin(e.start_time).strftime("%d.%m.%Y %H:%M") if e.start_time else "-",
                "end_time": as_berlin(e.end_time).strftime("%d.%m.%Y %H:%M") if e.end_time else "-",
                "regular_hours": f"{(reg_m / 60):.2f}",
                "overtime_hours": f"{(ot_m / 60):.2f}",
                "total_hours": f"{(tot_m / 60):.2f}",
                "total_cost_eur": eur(e.total_cost),
            }
        )
    corrections_log = db.scalars(
        select(TimeEntryCorrection)
        .where(
            or_(
                TimeEntryCorrection.old_employee_id == employee_id,
                TimeEntryCorrection.new_employee_id == employee_id,
            )
        )
        .order_by(desc(TimeEntryCorrection.created_at))
        .limit(100)
    ).all()
    correction_rows = [_correction_row_display(c) for c in corrections_log]
    return templates.TemplateResponse(
        request=request,
        name="admin_time_employee_profile.html",
        context={
            "request": request,
            "berlin_now": now_berlin().strftime("%d.%m.%Y %H:%M:%S"),
            "employee": employee,
            "worked_days": len(day_set),
            "total_normal_hours": f"{(total_reg_m / 60):.2f}",
            "total_overtime_hours": f"{(total_ot_m / 60):.2f}",
            "total_cost_eur": eur(total_cost),
            "hourly_normal_eur": eur(employee.hourly_rate),
            "overtime_rate_eur": eur(employee_overtime_rate(employee)),
            "history": history,
            "correction_rows": correction_rows,
        },
    )


@router.get("/admin-time/vehicles/{vehicle_id}/profile", response_class=HTMLResponse)
def admin_vehicle_profile(request: Request, vehicle_id: int, db: Session = Depends(get_db)):
    ensure_reporting_schema(db)
    vehicle = db.get(Vehicle, vehicle_id)
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    entries = db.scalars(
        select(TimeEntry)
        .where(TimeEntry.vehicle_id == vehicle_id, TimeEntry.status == "completed")
        .order_by(desc(TimeEntry.start_time))
    ).all()
    total_minutes = sum(int(e.total_minutes or 0) for e in entries)
    employee_names = sorted({e.employee_name for e in entries if e.employee_name})
    history: list[dict[str, object]] = []
    for e in entries:
        reg_m = int(e.regular_minutes or 0)
        ot_m = int(e.overtime_minutes or 0)
        tot_m = int(e.total_minutes or 0)
        history.append(
            {
                "entry_id": e.id,
                "employee_name": e.employee_name,
                "date": as_berlin(e.start_time).strftime("%Y-%m-%d") if e.start_time else "-",
                "start_time": as_berlin(e.start_time).strftime("%d.%m.%Y %H:%M") if e.start_time else "-",
                "end_time": as_berlin(e.end_time).strftime("%d.%m.%Y %H:%M") if e.end_time else "-",
                "regular_hours": f"{(reg_m / 60):.2f}",
                "overtime_hours": f"{(ot_m / 60):.2f}",
                "total_hours": f"{(tot_m / 60):.2f}",
                "total_cost_eur": eur(e.total_cost),
            }
        )
    corrections_log = db.scalars(
        select(TimeEntryCorrection)
        .where(
            or_(
                TimeEntryCorrection.old_vehicle_id == vehicle_id,
                TimeEntryCorrection.new_vehicle_id == vehicle_id,
            )
        )
        .order_by(desc(TimeEntryCorrection.created_at))
        .limit(100)
    ).all()
    correction_rows = [_correction_row_display(c) for c in corrections_log]
    return templates.TemplateResponse(
        request=request,
        name="admin_time_vehicle_profile.html",
        context={
            "request": request,
            "berlin_now": now_berlin().strftime("%d.%m.%Y %H:%M:%S"),
            "vehicle": vehicle,
            "time_url": f"{BASE_TIME_URL}?vehicle={vehicle.qr_code_slug}",
            "total_hours_used": round(total_minutes / 60.0, 2),
            "employee_names": employee_names,
            "history": history,
            "correction_rows": correction_rows,
        },
    )


@router.get("/admin-time/time-entries/{entry_id}/edit", response_class=HTMLResponse)
def admin_time_entry_edit_get(request: Request, entry_id: int, db: Session = Depends(get_db)):
    ensure_reporting_schema(db)
    entry = db.get(TimeEntry, entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Time entry not found")
    employees = db.scalars(select(Employee).order_by(Employee.name)).all()
    vehicles = db.scalars(select(Vehicle).order_by(Vehicle.name)).all()
    corrections = db.scalars(
        select(TimeEntryCorrection)
        .where(TimeEntryCorrection.time_entry_id == entry_id)
        .order_by(desc(TimeEntryCorrection.created_at))
        .limit(50)
    ).all()
    error = None
    if entry.status != "completed":
        error = "Bu kayıt henüz tamamlanmadı; manuel düzeltme yalnızca tamamlanmış mesailer için kullanılabilir."
    return templates.TemplateResponse(
        request=request,
        name="admin_time_time_entry_edit.html",
        context=_time_entry_edit_page_context(
            request,
            entry,
            employees,
            vehicles,
            corrections,
            clock_in_value=fmt_datetime_local(entry.start_time),
            clock_out_value=fmt_datetime_local(entry.end_time),
            normal_hours_value="",
            overtime_hours_value="",
            reason_value="",
            form_error=error,
        ),
    )


@router.post("/admin-time/time-entries/{entry_id}/edit", response_class=HTMLResponse)
def admin_time_entry_edit_post(
    request: Request,
    entry_id: int,
    employee_id: int = Form(),
    vehicle_id: int = Form(),
    clock_in: str = Form(""),
    clock_out: str = Form(""),
    normal_hours: str = Form(""),
    overtime_hours: str = Form(""),
    reason: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_reporting_schema(db)
    entry = db.get(TimeEntry, entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Time entry not found")
    if entry.status != "completed":
        return admin_time_entry_edit_get(request, entry_id, db)  # type: ignore[misc]

    employees = db.scalars(select(Employee).order_by(Employee.name)).all()
    vehicles = db.scalars(select(Vehicle).order_by(Vehicle.name)).all()
    corrections = db.scalars(
        select(TimeEntryCorrection)
        .where(TimeEntryCorrection.time_entry_id == entry_id)
        .order_by(desc(TimeEntryCorrection.created_at))
        .limit(50)
    ).all()

    reason_s = (reason or "").strip()
    if not reason_s:
        return templates.TemplateResponse(
            request=request,
            name="admin_time_time_entry_edit.html",
            context=_time_entry_edit_page_context(
                request,
                entry,
                employees,
                vehicles,
                corrections,
                clock_in_value=(clock_in or "").strip() or fmt_datetime_local(entry.start_time),
                clock_out_value=(clock_out or "").strip() or fmt_datetime_local(entry.end_time),
                normal_hours_value=(normal_hours or "").strip(),
                overtime_hours_value=(overtime_hours or "").strip(),
                reason_value=(reason or "").strip(),
                form_error="Düzeltme nedeni zorunludur.",
            ),
        )

    new_employee = db.get(Employee, employee_id)
    new_vehicle = db.get(Vehicle, vehicle_id)
    if not new_employee or not new_vehicle:
        return templates.TemplateResponse(
            request=request,
            name="admin_time_time_entry_edit.html",
            context=_time_entry_edit_page_context(
                request,
                entry,
                employees,
                vehicles,
                corrections,
                clock_in_value=(clock_in or "").strip() or fmt_datetime_local(entry.start_time),
                clock_out_value=(clock_out or "").strip() or fmt_datetime_local(entry.end_time),
                normal_hours_value=(normal_hours or "").strip(),
                overtime_hours_value=(overtime_hours or "").strip(),
                reason_value=reason_s,
                form_error="Geçerli çalışan ve araç seçin.",
            ),
        )

    nh = parse_optional_float_hours(normal_hours)
    oh = parse_optional_float_hours(overtime_hours)
    parsed_in = parse_datetime_local(clock_in)
    parsed_out = parse_datetime_local(clock_out)

    old_start = entry.start_time
    old_end = entry.end_time
    old_emp_id = entry.employee_id
    old_vehicle_id = entry.vehicle_id

    form_error: str | None = None
    new_start: datetime | None
    new_end: datetime | None
    total_m: int
    reg_m: int
    ot_m: int
    reg_c: float
    ot_c: float
    tot_c: float

    if nh is not None and oh is not None:
        if nh < 0 or oh < 0:
            form_error = "Saat değerleri negatif olamaz."
        else:
            new_start = parsed_in or entry.start_time
            new_end = parsed_out or entry.end_time
            total_m, reg_m, ot_m, reg_c, ot_c, tot_c = recalc_minutes_and_costs(
                new_employee,
                start=new_start,
                end=new_end,
                normal_hours_override=nh,
                overtime_hours_override=oh,
            )
    elif parsed_in and parsed_out:
        new_start, new_end = parsed_in, parsed_out
        total_m, reg_m, ot_m, reg_c, ot_c, tot_c = recalc_minutes_and_costs(
            new_employee,
            start=new_start,
            end=new_end,
            normal_hours_override=None,
            overtime_hours_override=None,
        )
    elif not (clock_in or "").strip() and not (clock_out or "").strip() and nh is None and oh is None:
        new_start = entry.start_time
        new_end = entry.end_time
        reg_m = int(entry.regular_minutes or 0)
        ot_m = int(entry.overtime_minutes or 0)
        total_m = reg_m + ot_m
        hr = float(new_employee.hourly_rate or 0)
        otr = employee_overtime_rate(new_employee)
        reg_c = round((reg_m / 60.0) * hr, 2)
        ot_c = round((ot_m / 60.0) * otr, 2)
        tot_c = round(reg_c + ot_c, 2)
    elif nh is not None or oh is not None:
        form_error = "Manuel saat için hem normal hem fazla mesai saatini girin (veya giriş/çıkış saatlerini kullanın)."
    else:
        form_error = "Giriş ve çıkış saatlerini birlikte girin veya saat alanlarını boş bırakıp yalnızca çalışan/araç değiştirin."

    if form_error:
        return templates.TemplateResponse(
            request=request,
            name="admin_time_time_entry_edit.html",
            context=_time_entry_edit_page_context(
                request,
                entry,
                employees,
                vehicles,
                corrections,
                clock_in_value=(clock_in or "").strip() or fmt_datetime_local(entry.start_time),
                clock_out_value=(clock_out or "").strip() or fmt_datetime_local(entry.end_time),
                normal_hours_value=(normal_hours or "").strip(),
                overtime_hours_value=(overtime_hours or "").strip(),
                reason_value=reason_s,
                form_error=form_error,
            ),
        )

    assert new_start is not None and new_end is not None

    by, role, client_ip, ua = correction_request_audit(request)
    correction = TimeEntryCorrection(
        time_entry_id=entry.id,
        old_clock_in=old_start,
        old_clock_out=old_end,
        new_clock_in=new_start,
        new_clock_out=new_end,
        old_employee_id=old_emp_id,
        new_employee_id=new_employee.id,
        old_vehicle_id=old_vehicle_id,
        new_vehicle_id=new_vehicle.id,
        reason=reason_s[:1000],
        created_at=now_berlin(),
        corrected_by=by,
        corrected_by_role=role,
        corrected_by_ip=client_ip,
        corrected_by_user_agent=ua,
    )
    db.add(correction)
    db.flush()
    entry.employee_id = new_employee.id
    entry.provisional_worker_id = None
    entry.employee_name = new_employee.name
    entry.vehicle_id = new_vehicle.id
    entry.start_time = new_start
    entry.end_time = new_end
    entry.total_minutes = total_m
    entry.regular_minutes = reg_m
    entry.overtime_minutes = ot_m
    entry.regular_cost = reg_c
    entry.overtime_cost = ot_c
    entry.total_cost = tot_c
    db.commit()
    return RedirectResponse("/admin-time/reports", status_code=302)


@router.get("/admin-time/reports/export")
def admin_time_reports_export(month: str, db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    try:
        year_s, month_s = month.split("-")
        year_i, month_i = int(year_s), int(month_s)
    except Exception:
        return Response(status_code=400, content=b"Invalid month format. Use YYYY-MM")

    employees = db.scalars(select(Employee).order_by(Employee.name)).all()
    employee_map = {e.id: e for e in employees}
    vehicles = db.scalars(select(Vehicle)).all()
    vehicle_map = {v.id: v for v in vehicles}
    entries_all = db.scalars(select(TimeEntry).where(TimeEntry.status == "completed")).all()
    entries = [
        e
        for e in entries_all
        if e.start_time and as_berlin(e.start_time).year == year_i and as_berlin(e.start_time).month == month_i
    ]

    employee_report: dict[int, dict[str, object]] = {}
    vehicle_report: dict[int, dict[str, object]] = {}
    daily_rows: list[list[object]] = []
    total_regular_minutes = 0
    total_overtime_minutes = 0
    total_payment = 0.0
    total_overtime_payment = 0.0

    for e in entries:
        emp = employee_map.get(e.employee_id) if e.employee_id is not None else None
        reg_min = int(e.regular_minutes or 0)
        ot_min = int(e.overtime_minutes or 0)
        total_min = int(e.total_minutes or 0)
        reg_cost = float(e.regular_cost or 0)
        ot_cost = float(e.overtime_cost or 0)
        tot_cost = float(e.total_cost or 0)
        if e.employee_id is not None:
            if e.employee_id not in employee_report:
                employee_report[e.employee_id] = {
                    "name": e.employee_name,
                    "phone": (emp.phone_number if emp else None) or "-",
                    "days": set(),
                    "reg_min": 0,
                    "ot_min": 0,
                    "total_min": 0,
                    "hourly_rate": float(emp.hourly_rate or 0) if emp else 0.0,
                    "overtime_hourly_rate": employee_overtime_rate(emp),
                    "reg_cost": 0.0,
                    "ot_cost": 0.0,
                    "total_cost": 0.0,
                    "missing_rate": float(emp.hourly_rate or 0) <= 0 if emp else True,
                }
            r = employee_report[e.employee_id]
            r["days"].add(as_berlin(e.start_time).strftime("%Y-%m-%d") if e.start_time else "-")
            r["reg_min"] += reg_min
            r["ot_min"] += ot_min
            r["total_min"] += total_min
            r["reg_cost"] += reg_cost
            r["ot_cost"] += ot_cost
            r["total_cost"] += tot_cost

        if e.vehicle_id not in vehicle_report:
            vehicle_report[e.vehicle_id] = {
                "name": vehicle_map[e.vehicle_id].name if e.vehicle_id in vehicle_map else f"ID {e.vehicle_id}",
                "type": (vehicle_map[e.vehicle_id].type if e.vehicle_id in vehicle_map else "-") or "-",
                "minutes": 0,
                "employees": set(),
                "cost": 0.0,
            }
        vr = vehicle_report[e.vehicle_id]
        vr["minutes"] += total_min
        if e.employee_id is not None:
            vr["employees"].add(e.employee_id)
        vr["cost"] += tot_cost

        total_regular_minutes += reg_min
        total_overtime_minutes += ot_min
        total_payment += tot_cost
        total_overtime_payment += ot_cost

        daily_rows.append(
            [
                as_berlin(e.start_time).strftime("%Y-%m-%d") if e.start_time else "-",
                e.employee_name,
                (emp.phone_number if emp else None) or "-",
                vehicle_map[e.vehicle_id].name if e.vehicle_id in vehicle_map else f"ID {e.vehicle_id}",
                as_berlin(e.start_time).strftime("%d.%m.%Y %H:%M") if e.start_time else "-",
                as_berlin(e.end_time).strftime("%d.%m.%Y %H:%M") if e.end_time else "-",
                round(reg_min / 60, 2),
                round(ot_min / 60, 2),
                round(total_min / 60, 2),
                round(reg_cost, 2),
                round(ot_cost, 2),
                round(tot_cost, 2),
            ]
        )

    wb = Workbook()
    ws_month = wb.active
    ws_month.title = "AYLIK_TOPLU_RAPOR"
    ws_daily = wb.create_sheet("GUNLUK_DETAY")
    ws_vehicle = wb.create_sheet("ARAC_BAZLI_OZET")
    ws_dashboard = wb.create_sheet("DASHBOARD")

    ws_month.append(
        [
            "Çalışan",
            "Telefon",
            "Ay",
            "Çalışılan gün",
            "Normal saat",
            "Fazla mesai saati",
            "Toplam saat",
            "Normal saat ücreti",
            "Fazla mesai saat ücreti",
            "Normal ücret",
            "Fazla mesai ücreti",
            "Toplam ödeme",
        ]
    )
    for row in sorted(employee_report.values(), key=lambda x: str(x["name"])):
        ws_month.append(
            [
                row["name"],
                row["phone"],
                month,
                len(row["days"]),
                round(row["reg_min"] / 60, 2),
                round(row["ot_min"] / 60, 2),
                round(row["total_min"] / 60, 2),
                round(row["hourly_rate"], 2),
                round(row["overtime_hourly_rate"], 2),
                round(row["reg_cost"], 2),
                round(row["ot_cost"], 2),
                round(row["total_cost"], 2),
            ]
        )
        if row["missing_rate"]:
            ws_month.append(["Ücret tanımlanmamış", row["name"]])

    ws_daily.append(
        [
            "Tarih",
            "Çalışan",
            "Telefon",
            "Araç/İş Makinesi",
            "Başlangıç",
            "Bitiş",
            "Normal saat",
            "Fazla mesai saati",
            "Toplam saat",
            "Normal ücret",
            "Fazla mesai ücreti",
            "Toplam ödeme",
        ]
    )
    for row in daily_rows:
        ws_daily.append(row)

    ws_vehicle.append(["Araç", "Araç tipi", "Toplam saat", "Çalışan sayısı", "Toplam işçilik maliyeti"])
    for row in sorted(vehicle_report.values(), key=lambda x: str(x["name"])):
        ws_vehicle.append(
            [
                row["name"],
                row["type"],
                round(row["minutes"] / 60, 2),
                len(row["employees"]),
                round(row["cost"], 2),
            ]
        )

    ws_dashboard.append(["KPI", "Değer"])
    ws_dashboard.append(["toplam çalışan", len(employee_report)])
    ws_dashboard.append(["toplam normal saat", round(total_regular_minutes / 60, 2)])
    ws_dashboard.append(["toplam fazla mesai saati", round(total_overtime_minutes / 60, 2)])
    ws_dashboard.append(["toplam ödeme", round(total_payment, 2)])
    ws_dashboard.append(["toplam fazla mesai ödemesi", round(total_overtime_payment, 2)])

    fill = PatternFill("solid", fgColor="1F4E79")
    white = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.auto_filter.ref = f"A1:{chr(64 + ws.max_column)}1"
        for cell in ws[1]:
            cell.fill = fill
            cell.font = white

    out = io.BytesIO()
    wb.save(out)
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=aylik_toplu_rapor_{month}.xlsx"},
    )


@router.get("/admin-time/export")
def export_xlsx(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    entries = db.scalars(select(TimeEntry).order_by(desc(TimeEntry.start_time))).all()
    employees = db.scalars(select(Employee).order_by(Employee.id)).all()
    devices = db.scalars(select(Device)).all()
    vehicles = db.scalars(select(Vehicle)).all()
    vehicle_map = {v.id: v.name for v in vehicles}

    wb = Workbook()
    sheets = {
        "DASHBOARD": wb.active,
        "PERSONEL": wb.create_sheet("PERSONEL"),
        "VERI_GIRISI": wb.create_sheet("VERI_GIRISI"),
        "HESAPLAMA": wb.create_sheet("HESAPLAMA"),
        "MALIYET_RAPORU": wb.create_sheet("MALIYET_RAPORU"),
        "ARAC_BAZLI_RAPOR": wb.create_sheet("ARAC_BAZLI_RAPOR"),
        "AYLIK_OZET": wb.create_sheet("AYLIK_OZET"),
    }
    sheets["DASHBOARD"].title = "DASHBOARD"

    fill = PatternFill("solid", fgColor="1F4E79")
    white = Font(color="FFFFFF", bold=True)

    total_minutes = sum(e.total_minutes or 0 for e in entries)
    total_overtime = sum(e.overtime_minutes or 0 for e in entries)
    total_cost = sum(float(e.total_cost or 0) for e in entries)
    total_overtime_cost = sum(float(e.overtime_cost or 0) for e in entries)
    active_count = sum(1 for e in entries if e.status == "active")
    ws = sheets["DASHBOARD"]
    ws.append(["KPI", "Değer"])
    ws.append(["Toplam Çalışan", len(employees)])
    ws.append(["Toplam Saat", round(total_minutes / 60, 2)])
    ws.append(["Fazla Mesai Saat", round(total_overtime / 60, 2)])
    ws.append(["Aktif Mesai", active_count])
    ws.append(["Toplam Maliyet (EUR)", round(total_cost, 2)])
    ws.append(["Toplam Fazla Mesai Maliyeti (EUR)", round(total_overtime_cost, 2)])

    for name, ws in sheets.items():
        if name == "DASHBOARD":
            continue
        if name == "PERSONEL":
            ws.append(
                [
                    "employee_id",
                    "employee_name",
                    "active",
                    "registered_device_count",
                    "hourly_rate_eur",
                    "overtime_hourly_rate_eur",
                    "total_hours",
                    "overtime_hours",
                    "total_cost_eur",
                ]
            )
            for emp in employees:
                device_count = sum(1 for d in devices if d.employee_id == emp.id and d.active)
                emp_entries = [e for e in entries if e.employee_id == emp.id]
                ws.append(
                    [
                        emp.id,
                        emp.name,
                        "yes" if emp.active else "no",
                        device_count,
                        float(emp.hourly_rate or 0),
                        employee_overtime_rate(emp),
                        round(sum((e.total_minutes or 0) for e in emp_entries) / 60, 2),
                        round(sum((e.overtime_minutes or 0) for e in emp_entries) / 60, 2),
                        round(sum(float(e.total_cost or 0) for e in emp_entries), 2),
                    ]
                )
        elif name == "VERI_GIRISI":
            ws.append(
                [
                    "tarih",
                    "çalışan",
                    "araç",
                    "başlangıç",
                    "bitiş",
                    "toplam dakika",
                    "normal dakika",
                    "fazla mesai",
                    "normal maliyet",
                    "fazla mesai maliyeti",
                    "toplam maliyet",
                    "durum",
                ]
            )
            for e in entries:
                ws.append(
                    [
                        str(e.start_time.date()) if e.start_time else "",
                        e.employee_name,
                        vehicle_map.get(e.vehicle_id, e.vehicle_id),
                        str(e.start_time or ""),
                        str(e.end_time or ""),
                        e.total_minutes or 0,
                        e.regular_minutes or 0,
                        e.overtime_minutes or 0,
                        round(float(e.regular_cost or 0), 2),
                        round(float(e.overtime_cost or 0), 2),
                        round(float(e.total_cost or 0), 2),
                        e.status,
                    ]
                )
        elif name == "HESAPLAMA":
            ws.append(["employee_name", "total_hours", "regular_hours", "overtime_hours", "total_cost"])
            for emp in employees:
                emp_entries = [e for e in entries if e.employee_id == emp.id]
                ws.append(
                    [
                        emp.name,
                        round(sum((e.total_minutes or 0) for e in emp_entries) / 60, 2),
                        round(sum((e.regular_minutes or 0) for e in emp_entries) / 60, 2),
                        round(sum((e.overtime_minutes or 0) for e in emp_entries) / 60, 2),
                        round(sum(float(e.total_cost or 0) for e in emp_entries), 2),
                    ]
                )
        elif name == "MALIYET_RAPORU":
            ws.append(
                [
                    "çalışan",
                    "normal saat",
                    "fazla mesai saati",
                    "saat ücreti",
                    "fazla mesai saat ücreti",
                    "normal maliyet",
                    "fazla mesai maliyeti",
                    "toplam maliyet",
                ]
            )
            for emp in employees:
                emp_entries = [e for e in entries if e.employee_id == emp.id and e.status == "completed"]
                ws.append(
                    [
                        emp.name,
                        round(sum((e.regular_minutes or 0) for e in emp_entries) / 60, 2),
                        round(sum((e.overtime_minutes or 0) for e in emp_entries) / 60, 2),
                        float(emp.hourly_rate or 0),
                        employee_overtime_rate(emp),
                        round(sum(float(e.regular_cost or 0) for e in emp_entries), 2),
                        round(sum(float(e.overtime_cost or 0) for e in emp_entries), 2),
                        round(sum(float(e.total_cost or 0) for e in emp_entries), 2),
                    ]
                )
        elif name == "ARAC_BAZLI_RAPOR":
            ws.append(["araç adı", "toplam saat", "çalışan sayısı", "toplam işçilik maliyeti"])
            grouped: dict[int, dict] = {}
            for e in entries:
                if e.vehicle_id not in grouped:
                    grouped[e.vehicle_id] = {"minutes": 0, "employees": set(), "cost": 0.0}
                grouped[e.vehicle_id]["minutes"] += int(e.total_minutes or 0)
                grouped[e.vehicle_id]["employees"].add(e.employee_id)
                grouped[e.vehicle_id]["cost"] += float(e.total_cost or 0)
            for vehicle_id, values in grouped.items():
                ws.append(
                    [
                        vehicle_map.get(vehicle_id, f"ID {vehicle_id}"),
                        round(values["minutes"] / 60, 2),
                        len(values["employees"]),
                        round(values["cost"], 2),
                    ]
                )
        elif name == "AYLIK_OZET":
            ws.append(["ay", "toplam saat", "normal saat", "fazla mesai saati", "toplam maliyet", "fazla mesai maliyeti"])
            monthly: dict[str, dict] = {}
            for e in entries:
                if not e.start_time:
                    continue
                key = as_berlin(e.start_time).strftime("%Y-%m")
                if key not in monthly:
                    monthly[key] = {"total": 0, "regular": 0, "overtime": 0, "cost": 0.0, "ot_cost": 0.0}
                monthly[key]["total"] += int(e.total_minutes or 0)
                monthly[key]["regular"] += int(e.regular_minutes or 0)
                monthly[key]["overtime"] += int(e.overtime_minutes or 0)
                monthly[key]["cost"] += float(e.total_cost or 0)
                monthly[key]["ot_cost"] += float(e.overtime_cost or 0)
            for key, values in sorted(monthly.items()):
                ws.append(
                    [
                        key,
                        round(values["total"] / 60, 2),
                        round(values["regular"] / 60, 2),
                        round(values["overtime"] / 60, 2),
                        round(values["cost"], 2),
                        round(values["ot_cost"], 2),
                    ]
                )

    for ws in sheets.values():
        ws.auto_filter.ref = f"A1:{chr(64 + ws.max_column)}1"
        for cell in ws[1]:
            cell.fill = fill
            cell.font = white

    out = io.BytesIO()
    wb.save(out)
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=qr_time_export.xlsx"},
    )


@router.get("/admin-time/vehicles", response_class=HTMLResponse)
def admin_time_vehicles(request: Request, message: str = "", db: Session = Depends(get_db)):
    vehicles = db.scalars(select(Vehicle).order_by(Vehicle.qr_code_slug)).all()
    active_vehicles = [v for v in vehicles if v.active]
    rows = []
    for v in active_vehicles:
        qr_link = f"{BASE_TIME_URL}?vehicle={v.qr_code_slug}"
        qr_img_url = f"/admin-time/vehicles/{v.qr_code_slug}/qr"
        rows.append(
            {
                "id": v.id,
                "name": v.name,
                "type": v.type or "-",
                "slug": v.qr_code_slug,
                "active": v.active,
                "qr_link": qr_link,
                "qr_img_url": qr_img_url,
                "qr_download_url": f"{qr_img_url}?printable=1",
            }
        )
    all_rows = [
        {
            "id": v.id,
            "name": v.name,
            "type": v.type or "",
            "slug": v.qr_code_slug,
            "active": v.active,
        }
        for v in vehicles
    ]
    return templates.TemplateResponse(
        request=request,
        name="admin_time_vehicles.html",
        context={"request": request, "vehicles": rows, "all_vehicles": all_rows, "message": message},
    )


@router.post("/admin-time/vehicles", response_class=HTMLResponse)
def create_vehicle(
    request: Request,
    name: str = Form(...),
    type: str = Form(""),
    qr_code_slug: str = Form(""),
    db: Session = Depends(get_db),
):
    clean_name = name.strip()
    clean_type = type.strip().lower() or None
    slug = to_slug(qr_code_slug) if qr_code_slug.strip() else to_slug(clean_name)
    if not clean_name:
        return admin_time_vehicles(request, message="Araç adı boş olamaz.", db=db)
    if not slug:
        return admin_time_vehicles(request, message="Slug üretilemedi.", db=db)
    exists = db.scalar(select(Vehicle).where(Vehicle.qr_code_slug == slug))
    if exists:
        return admin_time_vehicles(request, message="Bu slug zaten kayıtlı.", db=db)
    db.add(Vehicle(name=clean_name, type=clean_type, qr_code_slug=slug, active=True))
    db.commit()
    return admin_time_vehicles(request, message=f"{clean_name} eklendi.", db=db)


@router.post("/admin-time/vehicles/{vehicle_id}/update", response_class=HTMLResponse)
def update_vehicle(
    request: Request,
    vehicle_id: int,
    name: str = Form(...),
    type: str = Form(""),
    qr_code_slug: str = Form(...),
    db: Session = Depends(get_db),
):
    vehicle = db.scalar(select(Vehicle).where(Vehicle.id == vehicle_id))
    if not vehicle:
        return admin_time_vehicles(request, message="Araç bulunamadı.", db=db)
    clean_name = name.strip()
    clean_type = type.strip().lower() or None
    new_slug = to_slug(qr_code_slug)
    if not clean_name or not new_slug:
        return admin_time_vehicles(request, message="Araç adı ve slug zorunlu.", db=db)
    slug_owner = db.scalar(select(Vehicle).where(Vehicle.qr_code_slug == new_slug, Vehicle.id != vehicle_id))
    if slug_owner:
        return admin_time_vehicles(request, message="Slug başka araçta kullanılıyor.", db=db)
    vehicle.name = clean_name
    vehicle.type = clean_type
    vehicle.qr_code_slug = new_slug
    db.commit()
    return admin_time_vehicles(request, message=f"{clean_name} güncellendi.", db=db)


@router.post("/admin-time/vehicles/{vehicle_id}/deactivate", response_class=HTMLResponse)
def deactivate_vehicle(request: Request, vehicle_id: int, db: Session = Depends(get_db)):
    vehicle = db.scalar(select(Vehicle).where(Vehicle.id == vehicle_id))
    if not vehicle:
        return admin_time_vehicles(request, message="Araç bulunamadı.", db=db)
    vehicle.active = False
    db.commit()
    return admin_time_vehicles(request, message=f"{vehicle.name} pasif yapıldı.", db=db)


@router.post("/admin-time/vehicles/{vehicle_id}/activate", response_class=HTMLResponse)
def activate_vehicle(request: Request, vehicle_id: int, db: Session = Depends(get_db)):
    vehicle = db.scalar(select(Vehicle).where(Vehicle.id == vehicle_id))
    if not vehicle:
        return admin_time_vehicles(request, message="Araç bulunamadı.", db=db)
    vehicle.active = True
    db.commit()
    return admin_time_vehicles(request, message=f"{vehicle.name} tekrar aktif edildi.", db=db)


@router.get("/admin-time/vehicles/{vehicle_slug}/qr")
def vehicle_qr_png(vehicle_slug: str, printable: int = 0, db: Session = Depends(get_db)):
    vehicle = db.scalar(select(Vehicle).where(Vehicle.qr_code_slug == vehicle_slug, Vehicle.active.is_(True)))
    if not vehicle:
        return Response(status_code=404, content=b"Vehicle not found")
    target_url = f"{BASE_TIME_URL}?vehicle={vehicle.qr_code_slug}"
    if printable:
        from PIL import Image, ImageDraw, ImageFont

        qr = qrcode.QRCode(version=4, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=18, border=4)
        qr.add_data(target_url)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
        canvas = Image.new("RGB", (qr_img.width + 180, qr_img.height + 280), "white")
        canvas.paste(qr_img, (90, 50))
        draw = ImageDraw.Draw(canvas)
        font = ImageFont.load_default()
        label = (vehicle.name or "Bilinmiyor").upper()
        subtitle = f"VEHICLE ID: {vehicle.qr_code_slug}"
        draw.text((90, qr_img.height + 80), label, fill="black", font=font)
        draw.text((90, qr_img.height + 105), subtitle, fill="black", font=font)
        draw.text((90, qr_img.height + 130), "SCAN TO START WORK", fill="black", font=font)
        draw.rectangle([(60, 24), (canvas.width - 60, canvas.height - 24)], outline="black", width=2)
        img = canvas
    else:
        img = qrcode.make(target_url).convert("RGB")
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    return Response(content=buffer.getvalue(), media_type="image/png")
